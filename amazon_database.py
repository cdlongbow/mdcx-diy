"""
Amazon ASIN 数据库保存功能
用于保存影片番号与 ASIN 对应关系，方便后续统计和复用
"""

import asyncio
from datetime import datetime
from pathlib import Path
from typing import TypedDict


class AsinRecord(TypedDict, total=False):
    """ASIN 记录结构"""

    timestamp: str  # 记录时间
    number: str  # 影片番号
    asin: str  # 亚马逊 ASIN
    product_url: str  # 亚马逊商品详情页链接
    title: str  # 商品标题
    poster_url: str  # 封面图片 URL
    match_type: str  # 匹配类型 (barcode/number/actor 等)
    confidence: float  # 置信度
    search_keyword: str  # 搜索关键词
    media_type: str  # 媒体类型
    note: str  # 备注


async def save_asin_to_excel(
    records: list[AsinRecord],
    excel_path: Path | None = None,
    *,
    sheet_name: str = "ASIN 数据库",
    max_rows: int = 100000,
) -> Path:
    """
    保存 ASIN 记录到 Excel 文件

    Args:
        records: ASIN 记录列表
        excel_path: Excel 文件路径，默认保存到运行目录下的 amazon_asin_database.xlsx
        sheet_name: 工作表名称
        max_rows: 单个 sheet 最大行数，超过会自动分 sheet

    Returns:
        Excel 文件路径

    注意：
        需要安装 openpyxl 库：pip install openpyxl
    """
    try:
        import openpyxl
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise ImportError("请安装 openpyxl 库：pip install openpyxl")

    if excel_path is None:
        # 改进 3：默认保存到运行目录
        excel_path = Path.cwd() / "amazon_asin_database.xlsx"
    elif isinstance(excel_path, str):
        excel_path = Path(excel_path)

    excel_path.parent.mkdir(parents=True, exist_ok=True)

    try:
        wb = openpyxl.load_workbook(excel_path)
    except FileNotFoundError:
        wb = openpyxl.Workbook()

    ws = wb.active
    ws.title = sheet_name

    if not ws["A1"].value:
        headers = [
            "记录时间",
            "影片番号",
            "ASIN 编号",
            "影片链接",
            "商品标题",
            "封面 URL",
            "匹配类型",
            "置信度",
            "搜索关键词",
            "媒体类型",
            "备注",
        ]
        # 使用 cell() 直接设置表头，避免 append() 的空行问题
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = openpyxl.styles.Font(bold=True)
            cell.fill = openpyxl.styles.PatternFill("solid", fgColor="C0C0C0")
            cell.alignment = openpyxl.styles.Alignment(horizontal="center")

        ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"

    for record in records:
        row_data = [
            record.get("timestamp", datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
            record.get("number", ""),
            record.get("asin", ""),
            record.get("product_url", ""),
            record.get("title", ""),
            record.get("poster_url", ""),
            record.get("match_type", ""),
            record.get("confidence", 0.0),
            record.get("search_keyword", ""),
            record.get("media_type", ""),
            record.get("note", ""),
        ]
        ws.append(row_data)

    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)) + 2)
            except (TypeError, AttributeError):
                pass
        adjusted_width = min(max_length, 100)
        ws.column_dimensions[column].width = adjusted_width

    try:
        wb.save(excel_path)
        wb.close()
    except PermissionError as e:
        raise PermissionError(f"无法保存 Excel 文件，可能文件正被其他程序打开：{excel_path}") from e

    return excel_path


async def save_single_asin_record(
    number: str,
    asin: str,
    title: str = "",
    product_url: str = "",
    poster_url: str = "",
    match_type: str = "",
    confidence: float = 0.0,
    search_keyword: str = "",
    media_type: str = "",
    note: str = "",
    excel_path: Path | None = None,
) -> bool:
    """
    保存单条 ASIN 记录

    Args:
        number: 影片番号
        asin: ASIN 编号（必须为 10 位字母数字）
        title: 商品标题
        product_url: 亚马逊商品详情页链接
        poster_url: 封面图片 URL
        match_type: 匹配类型 (barcode/number/actor/soft 等)
        confidence: 置信度 (0-1 之间)
        search_keyword: 搜索关键词
        media_type: 媒体类型 (DVD/Blu-ray 等)
        note: 备注信息
        excel_path: Excel 文件路径

    Returns:
        bool: 保存成功返回 True，失败或跳过返回 False

    示例:
        success = await save_single_asin_record(
            number="ABC-123",
            asin="B0000001",
            title="作品标题",
            product_url="https://www.amazon.co.jp/dp/B0000001",
            match_type="number",
            confidence=0.95,
        )
        if success:
            print("保存成功")
        else:
            print("保存失败或跳过")
    """
    # 验证 ASIN
    import re
    
    if not asin or not asin.strip():
        return False
    
    asin = asin.strip().upper()
    
    # 验证 ASIN 格式（10 位字母数字）
    if not re.match(r"^[A-Z0-9]{10}$", asin):
        return False
    
    record: AsinRecord = {
        "number": number,
        "asin": asin,
        "product_url": product_url,
        "title": title,
        "poster_url": poster_url,
        "match_type": match_type,
        "confidence": confidence,
        "search_keyword": search_keyword,
        "media_type": media_type,
        "note": note,
    }

    try:
        await save_asin_to_excel([record], excel_path)
        return True
    except Exception:
        return False


async def query_asin_database(
    number: str | None = None,
    asin: str | None = None,
    excel_path: Path | None = None,
) -> list[AsinRecord]:
    """
    查询 ASIN 数据库

    Args:
        number: 按番号查询
        asin: 按 ASIN 查询
        excel_path: Excel 文件路径

    Returns:
        匹配的记录列表

    示例:
        results = await query_asin_database(number="ABC-123")
        results = await query_asin_database(asin="B0000001")
    """
    try:
        import openpyxl  # noqa: F401
    except ImportError:
        raise ImportError("请安装 openpyxl 库：pip install openpyxl")

    if excel_path is None:
        excel_path = Path.cwd() / "amazon_asin_database.xlsx"

    if not excel_path.exists():
        return []

    wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)
    ws = wb.active
    
    print(f"[DEBUG] Loaded file, max_row={ws.max_row}, max_col={ws.max_column}")

    results: list[AsinRecord] = []

    for row_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
        print(f"[DEBUG] Processing row {row_idx}: {len(row)} cols")
        if row_idx == 1:
            continue

        if len(row) < 11:
            continue

        record = AsinRecord(
            timestamp=str(row[0] or ""),
            number=str(row[1] or ""),
            asin=str(row[2] or ""),
            product_url=str(row[3] or ""),
            title=str(row[4] or ""),
            poster_url=str(row[5] or ""),
            match_type=str(row[6] or ""),
            confidence=float(row[7]) if row[7] else 0.0,
            search_keyword=str(row[8] or ""),
            media_type=str(row[9] or ""),
            note=str(row[10] or ""),
        )

        if number and str(record.get("number", "")).upper() == number.upper():
            results.append(record)
        elif asin and str(record.get("asin", "")).upper() == asin.upper():
            results.append(record)

    wb.close()
    return results


async def export_asin_statistics(
    excel_path: Path | None = None,
    output_path: Path | None = None,
) -> dict:
    """
    导出 ASIN 数据库统计信息

    Returns:
        统计信息字典
    """
    try:
        import openpyxl
    except ImportError:
        raise ImportError("请安装 openpyxl 库：pip install openpyxl")

    if excel_path is None:
        excel_path = Path.home() / ".mdcx" / "amazon_asin_database.xlsx"

    if not excel_path.exists():
        return {}

    if output_path is None:
        output_path = excel_path.parent / "amazon_statistics.txt"

    wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)
    ws = wb.active

    total_records = 0
    match_type_stats: dict[str, int] = {}
    media_type_stats: dict[str, int] = {}

    for row_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if row_idx == 1:
            continue

        if len(row) < 7:
            continue

        total_records += 1

        match_type = str(row[5] or "unknown")
        match_type_stats[match_type] = match_type_stats.get(match_type, 0) + 1

        media_type = str(row[8] or "unknown") if len(row) > 8 else "unknown"
        media_type_stats[media_type] = media_type_stats.get(media_type, 0) + 1

    wb.close()

    stats = {
        "total_records": total_records,
        "match_type_stats": match_type_stats,
        "media_type_stats": media_type_stats,
    }

    with open(output_path, "w", encoding="utf-8") as f:
        f.write("=" * 60 + "\n")
        f.write("Amazon ASIN 数据库统计报告\n")
        f.write(f"生成时间：{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
        f.write("=" * 60 + "\n\n")

        f.write(f"总记录数：{total_records}\n\n")

        f.write("匹配类型分布：\n")
        for match_type, count in sorted(match_type_stats.items(), key=lambda x: x[1], reverse=True):
            percentage = (count / total_records * 100) if total_records > 0 else 0
            f.write(f"  {match_type}: {count} 条 ({percentage:.1f}%)\n")

        f.write("\n媒体类型分布：\n")
        for media_type, count in sorted(media_type_stats.items(), key=lambda x: x[1], reverse=True):
            percentage = (count / total_records * 100) if total_records > 0 else 0
            f.write(f"  {media_type}: {count} 条 ({percentage:.1f}%)\n")

        f.write("\n" + "=" * 60 + "\n")

    return stats
