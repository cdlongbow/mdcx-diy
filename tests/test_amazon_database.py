import pytest
from openpyxl import load_workbook

from mdcx.core import amazon_database
from mdcx.models.log_buffer import LogBuffer

pytestmark = pytest.mark.network


@pytest.fixture
def _tmp_asin_db(monkeypatch: pytest.MonkeyPatch, tmp_path):
    from mdcx.config import manager

    monkeypatch.setattr(manager.manager, "data_folder", tmp_path)
    userdata = tmp_path / "userdata"
    userdata.mkdir(parents=True, exist_ok=True)
    return userdata / "amazon_asin_database.xlsx"


@pytest.mark.asyncio
async def test_save_single_asin_record_validates_asin_format(_tmp_asin_db):
    success = await amazon_database.save_single_asin_record(
        number="ABC-123",
        asin="invalid",
    )
    assert success is False


@pytest.mark.asyncio
async def test_save_single_asin_record_skips_empty_asin(_tmp_asin_db):
    success = await amazon_database.save_single_asin_record(
        number="ABC-123",
        asin="",
    )
    assert success is False


@pytest.mark.asyncio
async def test_save_single_asin_record_inserts_new_row(_tmp_asin_db):
    success = await amazon_database.save_single_asin_record(
        number="ABC-123",
        asin="B000000001",
        title="Test Title",
        product_url="https://www.amazon.co.jp/dp/B000000001",
        poster_url="https://m.media-amazon.com/images/I/test.jpg",
    )

    assert success is True

    wb = load_workbook(_tmp_asin_db)
    ws = wb.active
    assert ws.cell(row=2, column=1).value == "ABC-123"
    assert ws.cell(row=2, column=2).value == "B000000001"
    wb.close()


@pytest.mark.asyncio
async def test_save_asin_to_excel_formats_worksheet(_tmp_asin_db):
    records = [
        {
            "number": "ABC-123",
            "asin": "B000000001",
            "product_url": "https://www.amazon.co.jp/dp/B000000001",
            "poster_url": "https://m.media-amazon.com/images/I/test.jpg",
        }
    ]

    await amazon_database.save_asin_to_excel(records, _tmp_asin_db)

    wb = load_workbook(_tmp_asin_db)
    ws = wb.active

    assert ws.title == "ASIN 数据库"
    assert ws.freeze_panes == "A2"
    assert ws.auto_filter.ref == "A1:F2"

    headers = [ws.cell(row=1, column=i).value for i in range(1, 7)]
    assert headers == ["影片番号", "ASIN 编号", "影片链接", "商品标题", "封面 URL", "搜索关键词"]

    assert ws.cell(row=2, column=3).hyperlink.target == "https://www.amazon.co.jp/dp/B000000001"
    assert ws.cell(row=2, column=3).style == "Hyperlink"
    assert ws.cell(row=2, column=5).hyperlink.target == "https://m.media-amazon.com/images/I/test.jpg"
    assert ws.cell(row=2, column=5).style == "Hyperlink"

    wb.close()


@pytest.mark.asyncio
async def test_query_asin_database_by_number(_tmp_asin_db):
    await amazon_database.save_single_asin_record(
        number="ABC-123",
        asin="B000000001",
        title="Test Title",
    )

    results = await amazon_database.query_asin_database(number="ABC-123", excel_path=_tmp_asin_db)

    assert len(results) == 1
    assert results[0]["asin"] == "B000000001"


@pytest.mark.asyncio
async def test_query_asin_database_by_asin(_tmp_asin_db):
    await amazon_database.save_single_asin_record(
        number="ABC-123",
        asin="B000000001",
    )

    results = await amazon_database.query_asin_database(asin="B000000001", excel_path=_tmp_asin_db)

    assert len(results) == 1
    assert results[0]["number"] == "ABC-123"


@pytest.mark.asyncio
async def test_query_asin_database_returns_empty_when_file_not_exists():
    from pathlib import Path

    results = await amazon_database.query_asin_database(
        number="NONEXISTENT",
        excel_path=Path("/nonexistent/path/file.xlsx"),
    )

    assert results == []


@pytest.mark.asyncio
async def test_query_asin_database_logs_read_failure(monkeypatch, _tmp_asin_db):
    await amazon_database.save_single_asin_record(
        number="ABC-123",
        asin="B000000001",
    )

    import openpyxl

    def _raise_runtime_error(*args, **kwargs):
        raise RuntimeError("boom")

    monkeypatch.setattr(openpyxl, "load_workbook", _raise_runtime_error)
    LogBuffer.log().clear()

    results = await amazon_database.query_asin_database(number="ABC-123", excel_path=_tmp_asin_db)

    assert results == []
    assert "[ASIN 数据库] 读取失败：boom" in LogBuffer.log().get()


def test_format_asin_worksheet_logs_failure():
    LogBuffer.log().clear()

    amazon_database._format_asin_worksheet(object())

    assert "[ASIN 数据库] 工作表格式化失败：" in LogBuffer.log().get()


@pytest.mark.asyncio
async def test_save_asin_to_excel_logs_missing_openpyxl(monkeypatch, _tmp_asin_db):
    import sys

    original_openpyxl = sys.modules.get("openpyxl")
    sys.modules["openpyxl"] = None

    LogBuffer.log().clear()

    with pytest.raises(ImportError):
        await amazon_database.save_asin_to_excel([], _tmp_asin_db)

    assert "[ASIN 数据库] 缺少 openpyxl，无法保存 amazon_asin_database.xlsx" in LogBuffer.log().get()

    if original_openpyxl:
        sys.modules["openpyxl"] = original_openpyxl
    else:
        del sys.modules["openpyxl"]
