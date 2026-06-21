import json
from pathlib import Path

import pytest
from openpyxl import Workbook, load_workbook

from mdcx.core import tmdb_actor
from mdcx.models.log_buffer import LogBuffer


@pytest.fixture
def _tmp_actor_db(monkeypatch: pytest.MonkeyPatch, tmp_path: Path):
    monkeypatch.setattr(tmdb_actor.manager, "data_folder", tmp_path)
    monkeypatch.setattr(tmdb_actor.resources, "actor_db", {})
    userdata = tmp_path / "userdata"
    userdata.mkdir(parents=True, exist_ok=True)
    return userdata / "actor_database.xlsx"


@pytest.fixture(autouse=True)
def _reset_tmdb_query_cache():
    tmdb_actor._TMDB_QUERY_CACHE.clear()
    tmdb_actor._TMDB_QUERY_INFLIGHT.clear()


def test_expand_name_variants_supports_kana_ko_and_kanji_ko():
    assert tmdb_actor._expand_name_variants("加瀬かなこ") == {"加濑かなこ", "加濑かな子"}
    assert tmdb_actor._expand_name_variants("加瀬かな子") == {"加濑かなこ", "加濑かな子"}


def test_search_actor_db_reverse_matches_variant_and_alias(monkeypatch: pytest.MonkeyPatch):
    monkeypatch.setattr(
        tmdb_actor.resources,
        "actor_db",
        {
            "上原亜衣": {
                "zh_cn": "上原亚衣",
                "zh_tw": "上原亞衣",
                "keyword": "Rio,上原亚衣",
                "href": "",
                "tmdbid": 123,
                "tmdb_url": "https://www.themoviedb.org/person/123",
            }
        },
    )
    monkeypatch.setattr(tmdb_actor.resources, "actor_db_reverse_index", None)

    by_variant = tmdb_actor.search_actor_db_reverse("上原亚衣")
    by_alias = tmdb_actor.search_actor_db_reverse("Rio")

    assert by_variant is not None
    assert by_variant["jp"] == "上原亜衣"
    assert by_variant["tmdbid"] == 123

    assert by_alias is not None
    assert by_alias["jp"] == "上原亜衣"
    assert by_alias["tmdbid"] == 123


def test_search_actor_db_reverse_builds_reusable_index(monkeypatch: pytest.MonkeyPatch):
    monkeypatch.setattr(
        tmdb_actor.resources,
        "actor_db",
        {
            "加瀬かなこ": {
                "zh_cn": "加濑香奈子",
                "zh_tw": "加瀨香奈子",
                "keyword": "Kana,加濑かな子",
                "href": "",
                "tmdbid": 456,
                "tmdb_url": "https://www.themoviedb.org/person/456",
            }
        },
    )
    monkeypatch.setattr(tmdb_actor.resources, "actor_db_reverse_index", None)

    result = tmdb_actor.search_actor_db_reverse("加瀬かな子")

    assert result is not None
    assert result["jp"] == "加瀬かなこ"
    assert tmdb_actor.resources.actor_db_reverse_index is not None
    assert tmdb_actor.search_actor_db_reverse("Kana")["tmdbid"] == 456


def test_get_actor_data_uses_reverse_index_for_translated_name(monkeypatch: pytest.MonkeyPatch):
    monkeypatch.setattr(
        tmdb_actor.resources,
        "actor_db",
        {
            "上原亜衣": {
                "zh_cn": "上原亚衣",
                "zh_tw": "上原亞衣",
                "keyword": "Rio",
                "href": "https://example.com/uahara-ai",
                "tmdbid": 123,
                "tmdb_url": "https://www.themoviedb.org/person/123",
            }
        },
    )
    monkeypatch.setattr(tmdb_actor.resources, "actor_db_reverse_index", None)

    actor_data = tmdb_actor.resources.get_actor_data("上原亚衣")

    assert actor_data["has_name"] is True
    assert actor_data["jp"] == "上原亜衣"
    assert actor_data["zh_cn"] == "上原亚衣"
    assert actor_data["keyword"] == ["Rio"]
    assert actor_data["href"] == "https://example.com/uahara-ai"


def test_is_japan_place_supports_localized_place_names():
    assert tmdb_actor.is_japan_place("东京") is True
    assert tmdb_actor.is_japan_place("神奈川县") is True
    assert tmdb_actor.is_japan_place("中国广东") is False


@pytest.mark.asyncio
async def test_fetch_person_translations_distinguishes_cn_and_tw(monkeypatch: pytest.MonkeyPatch):
    class _StubResponse:
        def __init__(self, payload):
            self.status_code = 200
            self._payload = payload

        def json(self):
            return self._payload

    async def _stub_tmdb_request(*args, **kwargs):
        return _StubResponse(
            {
                "translations": [
                    {
                        "iso_639_1": "zh",
                        "iso_3166_1": "CN",
                        "english_name": "Kana",
                        "data": {"name": "加濑香奈"},
                    },
                    {
                        "iso_639_1": "zh",
                        "iso_3166_1": "TW",
                        "english_name": "Kana",
                        "data": {"name": "加瀨香奈"},
                    },
                ]
            }
        )

    monkeypatch.setattr(tmdb_actor, "_tmdb_request", _stub_tmdb_request)

    result = await tmdb_actor._fetch_person_translations(123, "https://api.tmdb.org", "token", object())

    assert result == {"zh_cn": "加濑香奈", "zh_tw": "加瀨香奈"}


@pytest.mark.asyncio
async def test_query_single_actor_tolerates_missing_http_response(monkeypatch: pytest.MonkeyPatch):
    async def _stub_tmdb_request(*args, **kwargs):
        return None

    monkeypatch.setattr(tmdb_actor, "_tmdb_request", _stub_tmdb_request)
    monkeypatch.setattr(tmdb_actor.manager.config, "show_data_log", False)

    result = await tmdb_actor._query_single_actor("上原亜衣", "https://api.tmdb.org", "token", object())

    assert result is None


@pytest.mark.asyncio
async def test_update_actor_db_row_allows_new_actor_insert_with_tmdbid(_tmp_actor_db: Path):
    status = await tmdb_actor.update_actor_db_row(
        jp="未知演员",
        zh_cn="未知演员",
        tmdbid=54321,
    )

    assert status == "inserted_new_row"

    wb = load_workbook(_tmp_actor_db)
    ws = wb.active
    assert ws.cell(row=2, column=1).value == "未知演员"
    assert ws.cell(row=2, column=6).value == 54321
    wb.close()


@pytest.mark.asyncio
async def test_update_actor_db_row_writes_tmdbid_and_tmdb_url(_tmp_actor_db: Path):
    status = await tmdb_actor.update_actor_db_row(
        jp="上原亜衣",
        zh_cn="上原亚衣",
        keyword="Rio",
        tmdbid=12345,
    )

    assert status == "inserted_new_row"

    wb = load_workbook(_tmp_actor_db)
    ws = wb.active
    assert ws.cell(row=2, column=1).value == "上原亜衣"
    assert ws.cell(row=2, column=2).value == "上原亚衣"
    assert ws.cell(row=2, column=6).value == 12345
    assert ws.cell(row=2, column=7).value == "https://www.themoviedb.org/person/12345"
    assert ws.cell(row=2, column=7).hyperlink.target == "https://www.themoviedb.org/person/12345"
    assert ws.title == "演员数据库"
    assert ws.freeze_panes == "B2"
    assert ws.auto_filter.ref == "A1:G2"
    wb.close()


def test_load_actor_db_derives_tmdb_url_from_tmdbid(_tmp_actor_db: Path):
    from openpyxl import load_workbook as lw

    wb = Workbook()
    ws = wb.active
    ws.append(tmdb_actor.DB_HEADERS)
    ws.append(["错位演员", "", "", "", "", 6233846, "https://www.themoviedb.org/person/6215799"])
    wb.save(_tmp_actor_db)
    wb.close()

    wb = lw(_tmp_actor_db, read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(min_row=2, max_row=2, values_only=True))
    row = rows[0]
    assert row[5] == 6233846
    assert row[6] == "https://www.themoviedb.org/person/6215799"
    wb.close()

    # 重新保存触发 _format_db_worksheet → tmdb_url 用 tmdbid 纠正
    wb2 = lw(_tmp_actor_db)
    ws2 = wb2.active
    from mdcx.core.tmdb_actor import _format_db_worksheet

    _format_db_worksheet(ws2)
    wb2.save(_tmp_actor_db)
    wb2.close()

    wb3 = lw(_tmp_actor_db)
    ws3 = wb3.active
    row3 = list(ws3.iter_rows(min_row=2, max_row=2, values_only=True))[0]
    assert row3[5] == 6233846
    assert row3[6] == f"https://www.themoviedb.org/person/{6233846}"


def test_tmdb_query_cache_path_fallback_to_userdata_when_invalid(monkeypatch: pytest.MonkeyPatch, tmp_path: Path):
    monkeypatch.setattr(tmdb_actor.manager, "data_folder", tmp_path)
    monkeypatch.setattr(tmdb_actor.resources, "tmdb_query_cache_path", Path(), raising=False)

    cache_path = tmdb_actor._tmdb_query_cache_path()
    assert cache_path == tmp_path / "userdata" / "actor_tmdb_query_cache.json"


def test_tmdb_query_cache_path_fallback_to_userdata_when_dot_path(monkeypatch: pytest.MonkeyPatch, tmp_path: Path):
    monkeypatch.setattr(tmdb_actor.manager, "data_folder", tmp_path)
    monkeypatch.setattr(tmdb_actor.resources, "tmdb_query_cache_path", Path("."), raising=False)

    cache_path = tmdb_actor._tmdb_query_cache_path()
    assert cache_path == tmp_path / "userdata" / "actor_tmdb_query_cache.json"


def test_tmdb_query_cache_sanitize_removes_stale_and_limits_size(monkeypatch: pytest.MonkeyPatch):
    now = 2_000_000_000.0
    monkeypatch.setattr(tmdb_actor, "_TMDB_QUERY_CACHE_MAX_ENTRIES", 8)

    tmdb_actor._TMDB_QUERY_CACHE.update(
        {
            "stale": (now - tmdb_actor._TMDB_QUERY_CACHE_TTL_SECONDS - 1, {"a": 1}),
            **{f"name-{i}": (now - i, {"idx": i}) for i in range(13)},
        }
    )

    tmdb_actor._tmdb_query_cache_sanitize(now=now)

    assert "stale" not in tmdb_actor._TMDB_QUERY_CACHE
    assert len(tmdb_actor._TMDB_QUERY_CACHE) == tmdb_actor._TMDB_QUERY_CACHE_MAX_ENTRIES
    for i in range(8):
        assert f"name-{i}" in tmdb_actor._TMDB_QUERY_CACHE
    for i in range(8, 13):
        assert f"name-{i}" not in tmdb_actor._TMDB_QUERY_CACHE


def test_tmdb_query_cache_persist_sync_filters_invalid_and_limits_size(monkeypatch: pytest.MonkeyPatch, tmp_path: Path):
    cache_file = tmp_path / "actor_tmdb_query_cache.json"
    monkeypatch.setattr(tmdb_actor.manager, "data_folder", tmp_path)
    monkeypatch.setattr(tmdb_actor.resources, "tmdb_query_cache_path", cache_file, raising=False)
    monkeypatch.setattr(tmdb_actor, "_TMDB_QUERY_CACHE_MAX_ENTRIES", 1, raising=False)
    now = 2_000_000_000.0

    tmdb_actor._TMDB_QUERY_CACHE.update(
        {
            "a": (now - 10, {"k": "v1"}),
            "b": (now - 5, {"k": "v2"}),
            "bad": (now - 1, "invalid"),
        }
    )

    tmdb_actor._tmdb_query_cache_persist_sync()

    loaded = json.loads(cache_file.read_text(encoding="utf-8"))
    assert list(loaded.keys()) == ["b"]


@pytest.mark.asyncio
async def test_update_actor_db_row_repairs_mismatched_tmdb_url(_tmp_actor_db: Path):
    wb = Workbook()
    ws = wb.active
    ws.append(tmdb_actor.DB_HEADERS)
    ws.append(["错位演员", "", "", "", "", 6233846, "https://www.themoviedb.org/person/6215799"])
    ws.cell(row=2, column=7).hyperlink = "https://www.themoviedb.org/person/6215799"
    wb.save(_tmp_actor_db)
    wb.close()

    status = await tmdb_actor.update_actor_db_row(jp="错位演员", tmdbid=6233846)

    assert status == "kept_existing_tmdbid"

    wb = load_workbook(_tmp_actor_db)
    ws = wb.active
    assert ws.cell(row=2, column=6).value == 6233846
    assert ws.cell(row=2, column=7).value == "https://www.themoviedb.org/person/6233846"
    assert ws.cell(row=2, column=7).hyperlink.target == "https://www.themoviedb.org/person/6233846"
    wb.close()


@pytest.mark.asyncio
async def test_update_actor_db_row_returns_inserted_tmdbid_for_existing_row(_tmp_actor_db: Path):
    await tmdb_actor.update_actor_db_row(jp="北條麻妃", zh_cn="北条麻妃")

    status = await tmdb_actor.update_actor_db_row(
        jp="北條麻妃",
        keyword="北条麻妃,Maki Hojo",
        tmdbid=67890,
        append_keyword=True,
    )

    assert status == "inserted_tmdbid"

    wb = load_workbook(_tmp_actor_db)
    ws = wb.active
    assert ws.cell(row=2, column=6).value == 67890
    assert ws.cell(row=2, column=7).value == "https://www.themoviedb.org/person/67890"
    keywords = {part.strip() for part in str(ws.cell(row=2, column=4).value or "").split(",") if part.strip()}
    assert keywords == {"北条麻妃", "Maki Hojo"}
    wb.close()


@pytest.mark.asyncio
async def test_update_actor_db_row_keeps_existing_tmdbid(_tmp_actor_db: Path):
    await tmdb_actor.update_actor_db_row(jp="加瀬かなこ", tmdbid=111)

    status = await tmdb_actor.update_actor_db_row(jp="加瀬かなこ", tmdbid=222)

    assert status == "kept_existing_tmdbid"

    wb = load_workbook(_tmp_actor_db)
    ws = wb.active
    assert ws.cell(row=2, column=6).value == 111
    assert ws.cell(row=2, column=7).value == "https://www.themoviedb.org/person/111"
    wb.close()


@pytest.mark.asyncio
async def test_update_actor_db_row_refreshes_resources_cache(_tmp_actor_db: Path):
    tmdb_actor.resources.actor_db = {}
    tmdb_actor.resources.actor_db_reverse_index = {"STALE": "旧数据"}

    status = await tmdb_actor.update_actor_db_row(jp="天使もえ", zh_cn="天使萌", tmdbid=24680)

    assert status == "inserted_new_row"
    assert tmdb_actor.resources.actor_db is not None
    assert tmdb_actor.resources.actor_db["天使もえ"]["tmdbid"] == 24680
    assert tmdb_actor.resources.actor_db_reverse_index is None


@pytest.mark.asyncio
async def test_update_actor_db_row_formats_headers_and_href_hyperlink(_tmp_actor_db: Path):
    status = await tmdb_actor.update_actor_db_row(
        jp="三上悠亜",
        zh_cn="三上悠亚",
        href="https://example.com/actor/mikami-yua",
    )

    assert status == "inserted_new_row"

    wb = load_workbook(_tmp_actor_db)
    ws = wb.active
    headers = [ws.cell(row=1, column=i).value for i in range(1, 8)]

    assert headers == tmdb_actor.DB_HEADERS
    assert ws.cell(row=2, column=5).hyperlink.target == "https://example.com/actor/mikami-yua"
    assert ws.cell(row=2, column=5).style == "Hyperlink"
    wb.close()


@pytest.mark.asyncio
async def test_update_actor_db_row_returns_file_locked_when_workbook_is_unavailable(
    _tmp_actor_db: Path, monkeypatch: pytest.MonkeyPatch
):
    await tmdb_actor.update_actor_db_row(jp="河北彩花", zh_cn="河北彩花")

    import openpyxl

    def _raise_permission_error(*args, **kwargs):
        raise PermissionError("locked")

    monkeypatch.setattr(openpyxl, "load_workbook", _raise_permission_error)

    status = await tmdb_actor.update_actor_db_row(jp="河北彩花", tmdbid=13579)

    assert status == "file_locked"


@pytest.mark.asyncio
async def test_load_actor_db_logs_read_failure(monkeypatch: pytest.MonkeyPatch, _tmp_actor_db: Path):
    await tmdb_actor.update_actor_db_row(jp="相泽南", zh_cn="相泽南")

    def _raise_runtime_error(*args, **kwargs):
        raise RuntimeError("boom")

    monkeypatch.setattr(tmdb_actor, "read_actor_db_xlsx", _raise_runtime_error)
    LogBuffer.log().clear()

    result = await tmdb_actor.load_actor_db()

    assert result == {}
    assert "[演员数据库] 读取失败: boom" in LogBuffer.log().get()


def test_format_db_worksheet_logs_failure_for_invalid_worksheet():
    LogBuffer.log().clear()

    tmdb_actor._format_db_worksheet(object())

    assert "[演员数据库] 工作表格式化失败:" in LogBuffer.log().get()


@pytest.mark.asyncio
async def test_update_actor_db_row_overwrite_names_overwrites_existing_zh_cn_zh_tw(_tmp_actor_db: Path):
    await tmdb_actor.update_actor_db_row(jp="波多野結衣", zh_cn="波多野结衣", zh_tw="波多野結衣", tmdbid=111)

    status = await tmdb_actor.update_actor_db_row(
        jp="波多野結衣", zh_cn="新译名", zh_tw="新繁體名", tmdbid=111, overwrite_names=True
    )

    assert status in ("updated_zh_cn", "updated_zh_tw", "updated_zh_cn_zh_tw")

    wb = load_workbook(_tmp_actor_db)
    ws = wb.active
    assert ws.cell(row=2, column=2).value == "新译名"
    assert ws.cell(row=2, column=3).value == "新繁體名"
    assert ws.cell(row=2, column=6).value == 111
    wb.close()


@pytest.mark.asyncio
async def test_update_actor_db_row_default_preserves_existing_zh_cn_zh_tw(_tmp_actor_db: Path):
    await tmdb_actor.update_actor_db_row(jp="波多野結衣", zh_cn="波多野结衣", zh_tw="波多野結衣", tmdbid=111)

    status = await tmdb_actor.update_actor_db_row(jp="波多野結衣", zh_cn="新译名", zh_tw="新繁體名", tmdbid=111)

    assert status in ("unchanged", "kept_existing_tmdbid")

    wb = load_workbook(_tmp_actor_db)
    ws = wb.active
    assert ws.cell(row=2, column=2).value == "波多野结衣"
    assert ws.cell(row=2, column=3).value == "波多野結衣"
    wb.close()


@pytest.mark.asyncio
async def test_update_actor_db_row_overwrite_fills_blank_fields(_tmp_actor_db: Path):
    await tmdb_actor.update_actor_db_row(jp="未知演员", zh_cn="", zh_tw="", tmdbid=222)

    status = await tmdb_actor.update_actor_db_row(
        jp="未知演员", zh_cn="新中文名", zh_tw="新繁體名", tmdbid=222, overwrite_names=True
    )

    assert status in ("updated_zh_cn", "updated_zh_tw", "updated_zh_cn_zh_tw")

    wb = load_workbook(_tmp_actor_db)
    ws = wb.active
    assert ws.cell(row=2, column=2).value == "新中文名"
    assert ws.cell(row=2, column=3).value == "新繁體名"
    wb.close()


@pytest.mark.asyncio
async def test_update_actor_db_row_overwrite_skips_empty_new_values(_tmp_actor_db: Path):
    await tmdb_actor.update_actor_db_row(jp="有名字的演员", zh_cn="原名", zh_tw="原名繁體", tmdbid=333)

    status = await tmdb_actor.update_actor_db_row(
        jp="有名字的演员", zh_cn="", zh_tw="", tmdbid=333, overwrite_names=True
    )

    assert status in ("unchanged", "kept_existing_tmdbid")

    wb = load_workbook(_tmp_actor_db)
    ws = wb.active
    assert ws.cell(row=2, column=2).value == "原名"
    assert ws.cell(row=2, column=3).value == "原名繁體"
    wb.close()


@pytest.mark.asyncio
async def test_fetch_actor_tmdb_ids_supplements_translation_for_cached_actor(
    monkeypatch: pytest.MonkeyPatch, _tmp_actor_db: Path
):
    await tmdb_actor.update_actor_db_row(jp="三上悠亜", zh_cn="", zh_tw="", tmdbid=12345)
    tmdb_actor.resources.actor_db = {
        "三上悠亜": {"zh_cn": "", "zh_tw": "", "keyword": "", "href": "", "tmdbid": 12345, "tmdb_url": ""}
    }
    tmdb_actor.resources.actor_db_reverse_index = None

    monkeypatch.setattr(tmdb_actor.manager.config, "tmdb_api_key", "fake-key")
    monkeypatch.setattr(tmdb_actor.manager.config, "tmdb_api_base", "api.tmdb.org")

    async def _stub_tmdb_request(client, method, url, **kwargs):
        if "/translations" in url:
            return tmdb_actor._TmdbResponse(
                200,
                '{"translations":[{"iso_639_1":"zh","iso_3166_1":"CN","english_name":"Mikami Yua","data":{"name":"三上悠亚"}},{"iso_639_1":"zh","iso_3166_1":"TW","english_name":"Mikami Yua","data":{"name":"三上悠亞"}}]}',
            )
        return None

    monkeypatch.setattr(tmdb_actor, "_tmdb_request", _stub_tmdb_request)

    result = await tmdb_actor.fetch_actor_tmdb_ids(["三上悠亜"], object())

    assert result == {"三上悠亜": 12345}

    wb = load_workbook(_tmp_actor_db)
    ws = wb.active
    for row in ws.iter_rows(min_row=2, values_only=True):
        if str(row[0] or "").strip() == "三上悠亜":
            assert row[1] == "三上悠亚"
            assert row[2] == "三上悠亞"
            break
    wb.close()


@pytest.mark.asyncio
async def test_fetch_actor_tmdb_ids_zhconv_fallback_for_new_actor(monkeypatch: pytest.MonkeyPatch, _tmp_actor_db: Path):
    monkeypatch.setattr(tmdb_actor.manager.config, "tmdb_api_key", "fake-key")
    monkeypatch.setattr(tmdb_actor.manager.config, "tmdb_api_base", "api.tmdb.org")

    async def _stub_tmdb_request(client, method, url, **kwargs):
        if "/search/person" in url:
            return tmdb_actor._TmdbResponse(
                200,
                '{"results":[{"id":99999,"name":"NewActor","original_name":"NewActor","gender":1,"adult":true,"popularity":1.0}]}',
            )
        if "/person/99999" in url and "/translations" not in url:
            return tmdb_actor._TmdbResponse(
                200,
                '{"id":99999,"name":"NewActor","original_name":"NewActor","gender":1,"place_of_birth":"Tokyo, Japan","also_known_as":["新俳優"]}',
            )
        if "/person/99999/translations" in url:
            return tmdb_actor._TmdbResponse(
                200,
                '{"translations":[{"iso_639_1":"zh","iso_3166_1":"CN","english_name":"","data":{"name":"新演员简体"}}]}',
            )
        return None

    monkeypatch.setattr(tmdb_actor, "_tmdb_request", _stub_tmdb_request)
    monkeypatch.setattr(tmdb_actor.manager.config, "show_data_log", False)

    result = await tmdb_actor.fetch_actor_tmdb_ids(["NewActor"], object())

    assert result == {"NewActor": 99999}

    wb = load_workbook(_tmp_actor_db)
    ws = wb.active
    found = False
    for row in ws.iter_rows(min_row=2, values_only=True):
        if str(row[0] or "").strip() == "NewActor":
            assert row[1] == "新演员简体"
            assert row[2] == "新演員簡體"
            found = True
            break
    assert found, "NewActor row not found in xlsx"
    wb.close()


@pytest.mark.asyncio
async def test_fetch_actor_tmdb_ids_keeps_site_name_as_jp_and_merges_tmdb_aliases(
    monkeypatch: pytest.MonkeyPatch, _tmp_actor_db: Path
):
    monkeypatch.setattr(tmdb_actor.manager.config, "tmdb_api_key", "fake-key")
    monkeypatch.setattr(tmdb_actor.manager.config, "tmdb_api_base", "api.tmdb.org")

    async def _stub_tmdb_request(client, method, url, **kwargs):
        if "/search/person" in url:
            return tmdb_actor._TmdbResponse(
                200,
                '{"results":[{"id":3714704,"name":"北条麻妃","original_name":"ほうじょう まき","gender":1,"adult":true,"popularity":1.0}]}',
            )
        if "/person/3714704" in url and "/translations" not in url:
            return tmdb_actor._TmdbResponse(
                200,
                '{"id":3714704,"name":"北条麻妃","original_name":"ほうじょう まき","gender":1,"place_of_birth":"Ishikawa, Japan","also_known_as":["Maki Houjou","ほうじょう まき","北条麻妃"]}',
            )
        if "/person/3714704/translations" in url:
            return tmdb_actor._TmdbResponse(
                200,
                '{"translations":[{"iso_639_1":"zh","iso_3166_1":"CN","english_name":"","data":{"name":"北条麻妃"}}]}',
            )
        return None

    monkeypatch.setattr(tmdb_actor, "_tmdb_request", _stub_tmdb_request)
    monkeypatch.setattr(tmdb_actor.manager.config, "show_data_log", False)

    result = await tmdb_actor.fetch_actor_tmdb_ids(["北條麻妃"], object())

    assert result == {"北條麻妃": 3714704}

    wb = load_workbook(_tmp_actor_db)
    ws = wb.active
    found = False
    for row in ws.iter_rows(min_row=2, values_only=True):
        if str(row[0] or "").strip() == "北條麻妃":
            assert row[1] == "北条麻妃"
            keywords = {part.strip() for part in str(row[3] or "").split(",") if part.strip()}
            assert keywords == {"北条麻妃", "ほうじょう まき", "Maki Houjou"}
            found = True
            break
    assert found, "北條麻妃 row not found in xlsx"
    wb.close()


@pytest.mark.asyncio
async def test_fetch_actor_tmdb_ids_skips_translate_when_both_names_present(
    monkeypatch: pytest.MonkeyPatch, _tmp_actor_db: Path
):
    await tmdb_actor.update_actor_db_row(jp="已完整演员", zh_cn="完整中文名", zh_tw="完整繁體名", tmdbid=55555)
    tmdb_actor.resources.actor_db = {
        "已完整演员": {
            "zh_cn": "完整中文名",
            "zh_tw": "完整繁體名",
            "keyword": "",
            "href": "",
            "tmdbid": 55555,
            "tmdb_url": "",
        }
    }
    tmdb_actor.resources.actor_db_reverse_index = None

    monkeypatch.setattr(tmdb_actor.manager.config, "tmdb_api_key", "fake-key")
    monkeypatch.setattr(tmdb_actor.manager.config, "tmdb_api_base", "api.tmdb.org")

    call_count = {"translations": 0}

    async def _stub_tmdb_request(client, method, url, **kwargs):
        if "/translations" in url:
            call_count["translations"] += 1
        return None

    monkeypatch.setattr(tmdb_actor, "_tmdb_request", _stub_tmdb_request)

    result = await tmdb_actor.fetch_actor_tmdb_ids(["已完整演员"], object())

    assert result == {"已完整演员": 55555}
    assert call_count["translations"] == 0
