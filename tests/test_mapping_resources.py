from mdcx.config.resources import resources


def test_get_info_data_matches_output_language_attrs(monkeypatch):
    monkeypatch.setattr(
        resources,
        "info_db",
        [{"zh_cn": "苗条", "zh_tw": "苗條", "jp": "スレンダー", "keyword": ",纖細,苗条的,苗條的,"}],
    )

    info_data = resources.get_info_data("スレンダー")

    assert info_data["has_name"] is True
    assert info_data["zh_cn"] == "苗条"
    assert info_data["zh_tw"] == "苗條"
    assert info_data["jp"] == "スレンダー"


def test_get_info_data_keeps_keyword_matching(monkeypatch):
    monkeypatch.setattr(
        resources,
        "info_db",
        [{"zh_cn": "足交", "zh_tw": "足交", "jp": "足交", "keyword": ",足コキ,足交,"}],
    )

    info_data = resources.get_info_data("足コキ")

    assert info_data["has_name"] is True
    assert info_data["zh_cn"] == "足交"


def test_reload_actor_db_keeps_existing_cache_on_failure(monkeypatch):
    old_actor_db = {
        "上原亜衣": {"zh_cn": "上原亚衣", "zh_tw": "上原亞衣", "keyword": "", "href": "", "tmdbid": 1, "tmdb_url": ""}
    }
    monkeypatch.setattr(resources, "actor_db", old_actor_db.copy())

    def _raise_load_workbook(*args, **kwargs):
        raise RuntimeError("boom")

    import openpyxl

    monkeypatch.setattr(openpyxl, "load_workbook", _raise_load_workbook)
    monkeypatch.setattr(resources, "u", lambda _name: type("_P", (), {"exists": lambda self: True})())

    resources.reload_actor_db()

    assert resources.actor_db == old_actor_db


def test_reload_actor_db_reads_shared_actor_xlsx_schema(monkeypatch, tmp_path):
    from openpyxl import Workbook

    userdata = tmp_path / "userdata"
    userdata.mkdir(parents=True, exist_ok=True)
    db_path = userdata / "actor_database.xlsx"

    wb = Workbook()
    ws = wb.active
    ws.append(["日文原名", "中文名", "繁体名", "别名", "链接", "tmdbid", "tmdb url"])
    ws.append(
        ["葵つかさ", "葵司", "葵司", "Aoi", "https://example.com/aoi", 9988, "https://www.themoviedb.org/person/9988"]
    )
    wb.save(db_path)
    wb.close()

    monkeypatch.setattr(resources, "u", lambda _name: db_path)
    monkeypatch.setattr(resources, "actor_db", None)
    monkeypatch.setattr(resources, "actor_db_reverse_index", {"OLD": "旧值"})

    resources.reload_actor_db()

    assert resources.actor_db == {
        "葵つかさ": {
            "zh_cn": "葵司",
            "zh_tw": "葵司",
            "keyword": "Aoi",
            "href": "https://example.com/aoi",
            "tmdbid": 9988,
            "tmdb_url": "https://www.themoviedb.org/person/9988",
        }
    }
    assert resources.actor_db_reverse_index is None
