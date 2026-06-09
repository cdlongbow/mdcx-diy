#!/usr/bin/env python3
"""
TMDB ID 验证工具 - 命令行简洁版
使用 TMDB 官方 API 验证演员数据库中的 TMDB ID

用法：
    python3 verify_simple.py <excel 文件路径> <API_KEY> [验证数量（可选）]

示例：
    python3 verify_simple.py resources/userdata/actor_database.xlsx YOUR_API_KEY 100
"""

import json
import sys
import time
import urllib.request
from datetime import datetime
from pathlib import Path

try:
    from openpyxl import load_workbook
    from openpyxl.styles import PatternFill
except ImportError:
    print("❌ 缺少 openpyxl：pip install openpyxl")
    sys.exit(1)

try:
    import certifi

    HAS_CERTIFI = True
except ImportError:
    HAS_CERTIFI = False

# ==================== 配置 ====================

TMDB_API_BASE = "https://api.themoviedb.org/3"
SAVE_INTERVAL = 200
COL_STATUS = 8  # H 列
COL_RESULT = 9  # I 列
COL_DATE = 10  # J 列

# ==================== 验证函数 ====================


def verify_with_api(tmdbid: int, api_key: str) -> tuple[bool, str]:
    """使用 TMDB API 验证演员 ID"""
    try:
        url = f"{TMDB_API_BASE}/person/{tmdbid}"
        params = f"?api_key={api_key}&language=ja-JP"

        req = urllib.request.Request(url + params)
        if HAS_CERTIFI:
            with urllib.request.urlopen(req, timeout=10, cafile=certifi.where()) as resp:
                data = json.loads(resp.read().decode("utf-8"))
        else:
            with urllib.request.urlopen(req, timeout=10) as resp:
                data = json.loads(resp.read().decode("utf-8"))

        name = data.get("name", "")
        aka = data.get("also_known_as", []) or []

        return True, name, aka

    except urllib.error.HTTPError as e:
        return False, f"HTTP {e.code}", []
    except Exception as e:
        return False, str(e), []


def main():
    if len(sys.argv) < 3:
        print(__doc__)
        sys.exit(1)

    excel_path = Path(sys.argv[1])
    api_key = sys.argv[2].strip()
    sample_size = int(sys.argv[3]) if len(sys.argv) > 3 else 0

    if not excel_path.exists():
        print(f"❌ 文件不存在：{excel_path}")
        sys.exit(1)

    # 加载 Excel
    print(f"📂 加载文件：{excel_path}")
    wb = load_workbook(excel_path)
    ws = wb.active

    # 添加状态列（如果不存在）
    if ws.cell(row=1, column=COL_STATUS).value != "验证状态":
        ws.cell(row=1, column=COL_STATUS, value="验证状态")
        ws.cell(row=1, column=COL_RESULT, value="验证结果")
        ws.cell(row=1, column=COL_DATE, value="验证日期")
        ws.column_dimensions["H"].width = 12
        ws.column_dimensions["I"].width = 50
        ws.column_dimensions["J"].width = 15
        wb.save(excel_path)
        print("✅ 已添加状态列")

    # 重新加载
    wb.close()
    wb = load_workbook(excel_path)
    ws = wb.active

    # 收集待验证记录
    to_verify = []
    for row in range(2, ws.max_row + 1):
        tmdbid = ws.cell(row=row, column=6).value
        tmdb_url = str(ws.cell(row=row, column=7).value or "").strip()
        jp_name = str(ws.cell(row=row, column=1).value or "").strip()
        status = str(ws.cell(row=row, column=COL_STATUS).value or "").strip()

        # 跳过已验证的
        if status in ["✅ 匹配", "❌ 不匹配"]:
            continue

        if tmdbid and tmdb_url.startswith("https://www.themoviedb.org/person/"):
            try:
                url_id = int(tmdb_url.split("/")[-1])
                if url_id == tmdbid:
                    to_verify.append((row, jp_name, tmdbid))
            except ValueError:
                pass

    if sample_size > 0:
        to_verify = to_verify[:sample_size]

    print(f"📋 找到 {len(to_verify)} 条待验证记录")
    if sample_size > 0:
        print(f"   (抽样 {sample_size} 条)")
    print()

    if not to_verify:
        print("✅ 所有记录已验证完毕")
        sys.exit(0)

    # 开始验证
    print("开始验证...\n")
    start_time = time.time()
    stats = {"matched": 0, "mismatch": 0, "error": 0}
    mismatches = []

    for i, (row, jp_name, tmdbid) in enumerate(to_verify, 1):
        # API 验证
        success, result, aka = verify_with_api(tmdbid, api_key)

        if not success:
            status = "🔴 错误"
            result_text = result
            stats["error"] += 1
        else:
            # 检查匹配
            name_match = jp_name == result or jp_name in aka

            if name_match:
                status = "✅ 匹配"
                result_text = result
                stats["matched"] += 1
            else:
                status = "❌ 不匹配"
                aka_preview = ", ".join(aka[:3]) if aka else "无别名"
                result_text = f"TMDB:{result} (别名:{aka_preview}...)"
                stats["mismatch"] += 1
                mismatches.append((row, jp_name, tmdbid, result_text))

        # 更新 Excel
        ws.cell(row=row, column=COL_STATUS, value=status)
        ws.cell(row=row, column=COL_RESULT, value=result_text)
        ws.cell(row=row, column=COL_DATE, value=datetime.now().strftime("%Y-%m-%d %H:%M"))

        # 颜色标记
        cell = ws.cell(row=row, column=COL_STATUS)
        if "✅" in status:
            cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        elif "❌" in status:
            cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        elif "🔴" in status:
            cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

        # 定期保存
        if i % SAVE_INTERVAL == 0:
            wb.save(excel_path)
            print(f"  💾 已保存进度 ({i}/{len(to_verify)})")

        # 进度显示
        elapsed = time.time() - start_time
        eta = ""
        if i > 0:
            remaining = len(to_verify) - i
            eta_sec = (elapsed / i) * remaining
            eta = f" ETA:{int(eta_sec / 60)}m{int(eta_sec % 60)}s"

        print(f"[{i}/{len(to_verify)}] {jp_name} -> {result_text}{eta}")

        # 限流（TMDB API 限制：每秒 3-4 次）
        time.sleep(0.35)

    # 最终保存
    wb.save(excel_path)
    elapsed = time.time() - start_time
    print(f"\n✅ 验证完成！用时：{int(elapsed / 60)}m{int(elapsed % 60)}s")
    print(f"   总计：{len(to_verify)} 条")
    print(f"   ✅ 匹配：{stats['matched']}")
    print(f"   ❌ 不匹配：{stats['mismatch']}")
    print(f"   🔴 错误：{stats['error']}")

    # 生成报告
    report_path = excel_path.parent / f"tmdb_verification_{datetime.now().strftime('%Y%m%d_%H%M')}.txt"
    with open(report_path, "w", encoding="utf-8") as f:
        f.write("TMDB ID 验证报告\n")
        f.write("=" * 70 + "\n\n")
        f.write(f"验证时间：{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
        f.write(f"Excel 文件：{excel_path}\n")
        f.write(f"验证总数：{len(to_verify)}\n\n")
        f.write("统计:\n")
        f.write(f"  ✅ 匹配：{stats['matched']}\n")
        f.write(f"  ❌ 不匹配：{stats['mismatch']}\n")
        f.write(f"  🔴 错误：{stats['error']}\n\n")

        if mismatches:
            f.write(f"\n❌ 不匹配记录（{len(mismatches)}条）:\n")
            f.write("-" * 70 + "\n")
            for row, jp, tid, result in mismatches:
                f.write(f"行{row}: {jp} -> TMDB ID:{tid} ({result})\n")

    print(f"\n📊 报告已生成：{report_path}")
    print("\n不匹配的记录已保存在 Excel 的 H 列（标记为❌）和上面的报告中")


if __name__ == "__main__":
    main()
