#!/usr/bin/env python3
"""
验证演员数据库中 TMDB ID 的正确性。
解析 TMDB 网页标题检查演员名字是否匹配。
"""

import re
import sys
import urllib.error
import urllib.request
from pathlib import Path

from openpyxl import load_workbook


def verify_tmdb_ids(excel_path: str, sample_size: int = 50):
    """验证 Excel 中的 TMDB ID 是否正确"""

    wb = load_workbook(excel_path, read_only=True)
    ws = wb.active

    COL_JP = 1
    COL_TMDBID = 6
    COL_TMDB_URL = 7

    to_verify = []
    for row_idx in range(2, ws.max_row + 1):
        tmdbid = ws.cell(row=row_idx, column=COL_TMDBID).value
        tmdb_url = str(ws.cell(row=row_idx, column=COL_TMDB_URL).value or "").strip()
        jp_name = str(ws.cell(row=row_idx, column=COL_JP).value or "").strip()

        if tmdbid and tmdb_url.startswith("https://www.themoviedb.org/person/"):
            try:
                url_id = int(tmdb_url.split("/")[-1])
                if url_id == tmdbid:
                    to_verify.append((jp_name, tmdbid, row_idx))
            except ValueError:
                pass

    wb.close()

    print(f"📋 从 URL 提取 tmdbid 的记录：{len(to_verify)} 条")
    print(f"🔍 本次验证：{sample_size} 条\n")

    import random

    samples = random.sample(to_verify, min(sample_size, len(to_verify)))
    samples.sort(key=lambda x: x[2])

    match_count = 0
    mismatch_count = 0
    error_count = 0
    mismatches = []

    for jp_name, tmdbid, row_idx in samples:
        try:
            url = f"https://www.themoviedb.org/person/{tmdbid}"
            with urllib.request.urlopen(url, timeout=8) as resp:
                html = resp.read().decode("utf-8")
                # 提取标题
                match = re.search(r"<title>([^-]+)", html)
                if match:
                    tmdb_name = match.group(1).strip()
                    # 清理 HTML 实体
                    tmdb_name = re.sub(r"&#\d+;?", lambda m: chr(int(m.group(0)[2:-1])), tmdb_name)

                    # 检查匹配
                    name_match = jp_name == tmdb_name

                    if name_match:
                        match_count += 1
                        print(f"✅ 行{row_idx}: {jp_name}")
                    else:
                        mismatch_count += 1
                        mismatches.append((row_idx, jp_name, tmdbid, tmdb_name))
                        print(f"❌ 行{row_idx}: {jp_name} <-> {tmdb_name}")
        except urllib.error.HTTPError as e:
            error_count += 1
            print(f"⚠️  行{row_idx}: HTTP {e.code}")
        except urllib.error.URLError:
            error_count += 1
            print(f"⚠️  行{row_idx}: URL 错误")
        except Exception as e:
            error_count += 1
            print(f"⚠️  行{row_idx}: {type(e).__name__}")

    # 报告
    print("\n" + "=" * 70)
    print("📊 验证报告")
    print("=" * 70)
    total = match_count + mismatch_count + error_count
    if total > 0:
        print(f"总验证数：{total}")
        print(f"✅ 匹配：{match_count} ({match_count / total * 100:.1f}%)")
        print(f"❌ 不匹配：{mismatch_count}")
        print(f"⚠️  错误：{error_count}")

    if mismatches:
        print(f"\n🔴 不匹配记录详情（共{len(mismatches)}条）:")
        for row, jp, tid, tmdb in mismatches[:20]:
            print(f"  行{row}: {jp} -> TMDB ID:{tid} (TMDB 名:{tmdb})")

        print("\n💡 建议：这些记录的 tmdbid 可能需要修正或删除")

    return mismatch_count == 0


if __name__ == "__main__":
    excel_path = Path(__file__).parent.parent / "resources" / "userdata" / "actor_database.xlsx"
    sample_size = int(sys.argv[1]) if len(sys.argv) > 1 else 50
    success = verify_tmdb_ids(str(excel_path), sample_size)
    sys.exit(0 if success else 1)
