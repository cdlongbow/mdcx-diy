#!/usr/bin/env python3
"""
演员数据库 TMDB ID 验证工具（GUI 版）

功能：
- 带 UI 界面，实时显示验证进度
- 断点续传：验证中断后可继续
- 每 200 条自动保存一次
- 自动在 Excel 中标记验证状态
- 生成详细报告便于后续处理
"""

import re
import sys
import urllib.error
import urllib.request
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook

# ==================== 检查依赖 ====================

try:
    import tkinter as tk
    from tkinter import filedialog, messagebox, scrolledtext, ttk
except ImportError:
    print("❌ 缺少 tkinter 库（Python 内置，但可能需要安装 python3-tk）")
    sys.exit(1)

try:
    import certifi

    HAS_CERTIFI = True
except ImportError:
    HAS_CERTIFI = False

# ==================== 验证逻辑 ====================


class TMDBVerifier:
    """TMDB ID 验证器"""

    def __init__(self, excel_path: str):
        self.excel_path = Path(excel_path)
        self.wb = None
        self.ws = None
        self.to_verify = []
        self.results = []
        self.stats = {
            "verified": 0,
            "matched": 0,
            "mismatch": 0,
            "error": 0,
            "no_network": 0,
        }

        # 验证状态列索引
        self.COL_STATUS = 8  # H 列：验证状态
        self.COL_RESULT = 9  # I 列：验证结果
        self.COL_CHECKED = 10  # J 列：验证日期

        # 添加状态列
        self._add_status_columns()

    def _add_status_columns(self):
        """添加验证状态列"""
        self.wb = load_workbook(self.excel_path)
        self.ws = self.wb.active

        # 检查是否已有状态列
        header_row = self.ws.cell(row=1, column=self.COL_STATUS)
        if header_row.value != "验证状态":
            # 添加表头
            self.ws.cell(row=1, column=self.COL_STATUS, value="验证状态")
            self.ws.cell(row=1, column=self.COL_RESULT, value="验证结果")
            self.ws.cell(row=1, column=self.COL_CHECKED, value="验证日期")

            # 调整列宽
            self.ws.column_dimensions["H"].width = 12
            self.ws.column_dimensions["I"].width = 50
            self.ws.column_dimensions["J"].width = 15

            self.wb.save(self.excel_path)

        # 重新加载（确保读取到最新数据）
        self.wb.close()
        self.wb = load_workbook(self.excel_path)
        self.ws = self.wb.active

    def find_records_to_verify(self):
        """找出需要验证的记录"""
        self.wb = load_workbook(self.excel_path, read_only=True)
        self.ws = self.wb.active

        COL_JP = 1
        COL_TMDBID = 6
        COL_TMDB_URL = 7

        self.to_verify = []

        for row_idx in range(2, self.ws.max_row + 1):
            tmdbid = self.ws.cell(row=row_idx, column=COL_TMDBID).value
            tmdb_url = str(self.ws.cell(row=row_idx, column=COL_TMDB_URL).value or "").strip()
            jp_name = str(self.ws.cell(row=row_idx, column=COL_JP).value or "").strip()
            status = str(self.ws.cell(row=row_idx, column=self.COL_STATUS).value or "").strip()

            # 只验证未验证的记录
            if status in ["✅ 匹配", "❌ 不匹配", "⚠️ 无网络"]:
                continue

            if tmdbid and tmdb_url.startswith("https://www.themoviedb.org/person/"):
                try:
                    url_id = int(tmdb_url.split("/")[-1])
                    if url_id == tmdbid:
                        self.to_verify.append((row_idx, jp_name, tmdbid))
                except ValueError:
                    pass

        self.wb.close()

        # 重新加载为非只读模式
        self.wb = load_workbook(self.excel_path)
        self.ws = self.wb.active

        return len(self.to_verify)

    def verify_single(self, row_idx: int, jp_name: str, tmdbid: int):
        """验证单条记录"""
        try:
            url = f"https://www.themoviedb.org/person/{tmdbid}"
            req = urllib.request.Request(url, headers={"User-Agent": "Mozilla/5.0"})

            if HAS_CERTIFI:
                with urllib.request.urlopen(req, timeout=10, cafile=certifi.where()) as resp:
                    html = resp.read().decode("utf-8")
            else:
                with urllib.request.urlopen(req, timeout=10) as resp:
                    html = resp.read().decode("utf-8")

            # 提取标题
            match = re.search(r"<title>([^-]+)", html)
            if match:
                tmdb_name = match.group(1).strip()
                # 清理 HTML 实体和多余内容
                tmdb_name = re.sub(r"\s*&#8212;.*$", "", tmdb_name)
                tmdb_name = re.sub(r"&#(\d+);", lambda m: chr(int(m.group(1))), tmdb_name)

                # 检查匹配
                name_match = jp_name == tmdb_name or jp_name in html

                if name_match:
                    return "matched", f"✅ {tmdb_name}"
                else:
                    return "mismatch", f"❌ TMDB: {tmdb_name}"
            else:
                return "error", "无法解析页面"

        except urllib.error.HTTPError as e:
            if e.code == 404:
                return "mismatch", "❌ TMDB ID 不存在"
            return "error", f"HTTP {e.code}"
        except urllib.error.URLError as e:
            if "Network" in str(e) or "Temporary" in str(e):
                return "no_network", "⚠️ 无网络"
            return "error", "URL 错误"
        except Exception as e:
            return "error", f"{type(e).__name__}"

    def save_progress(self, count: int = 0):
        """保存进度（每 200 条或结束时调用）"""
        if self.wb:
            self.wb.save(self.excel_path)
            print(f"💾 已保存进度 (已验证：{count} 条)")

    def update_cell(self, row_idx: int, status: str, result: str):
        """更新单元格"""
        self.ws.cell(row=row_idx, column=self.COL_STATUS, value=status)
        self.ws.cell(row=row_idx, column=self.COL_RESULT, value=result)
        self.ws.cell(row=row_idx, column=self.COL_CHECKED, value=datetime.now().strftime("%Y-%m-%d %H:%M"))

        # 根据状态设置颜色
        cell = self.ws.cell(row=row_idx, column=self.COL_STATUS)
        if status.startswith("✅"):
            cell.fill = openpyxl.styles.PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        elif status.startswith("❌"):
            cell.fill = openpyxl.styles.PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        elif status.startswith("⚠️"):
            cell.fill = openpyxl.styles.PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")

    def generate_report(self):
        """生成验证报告"""
        report_path = self.excel_path.parent / f"tmdb_verification_report_{datetime.now().strftime('%Y%m%d_%H%M')}.txt"

        mismatches = [(r[0], r[1], r[2], r[3]) for r in self.results if r[2] == "mismatch"]
        errors = [(r[0], r[1], r[2], r[3]) for r in self.results if r[2] == "error"]

        with open(report_path, "w", encoding="utf-8") as f:
            f.write("TMDB ID 验证报告\n")
            f.write("=" * 70 + "\n\n")
            f.write(f"验证时间：{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
            f.write(f"Excel 文件：{self.excel_path}\n\n")

            f.write("统计:\n")
            f.write(f"  总验证数：{len(self.results)}\n")
            f.write(f"  ✅ 匹配：{self.stats['matched']}\n")
            f.write(f"  ❌ 不匹配：{self.stats['mismatch']}\n")
            f.write(f"  ⚠️  无网络：{self.stats['no_network']}\n")
            f.write(f"  🔴 错误：{self.stats['error']}\n\n")

            if mismatches:
                f.write(f"\n❌ 不匹配记录（{len(mismatches)}条）:\n")
                f.write("-" * 70 + "\n")
                for row, jp, tid, result in mismatches[:100]:
                    f.write(f"行{row}: {jp} -> TMDB ID:{tid} ({result.replace('❌ ', '')})\n")

            if errors:
                f.write(f"\n🔴 错误记录（{len(errors)}条）:\n")
                f.write("-" * 70 + "\n")
                for row, jp, tid, result in errors[:100]:
                    f.write(f"行{row}: {jp} -> TMDB ID:{tid} ({result})\n")

        return report_path


try:
    import openpyxl
except ImportError:
    print("❌ 缺少 openpyxl 库，请安装：pip install openpyxl")
    sys.exit(1)


# ==================== GUI 界面 ====================


class VerifierGUI:
    """验证工具 GUI"""

    def __init__(self, root):
        self.root = root
        self.root.title("🔍 TMDB ID 验证工具")
        self.root.geometry("900x650")

        self.verifier = None
        self.is_running = False
        self.save_interval = 200  # 每 200 条保存一次

        self._create_ui()

    def _create_ui(self):
        """创建 UI"""
        # 顶部：文件选择
        file_frame = ttk.LabelFrame(self.root, text="1. 选择 Excel 文件", padding=10)
        file_frame.pack(fill="x", padx=10, pady=5)

        self.file_var = tk.StringVar()
        file_entry = ttk.Entry(file_frame, textvariable=self.file_var, width=80)
        file_entry.pack(side="left", fill="x", expand=True, padx=(0, 10))

        browse_btn = ttk.Button(file_frame, text="浏览...", command=self._browse_file)
        browse_btn.pack(side="left")

        load_btn = ttk.Button(file_frame, text="加载文件", command=self._load_file)
        load_btn.pack(side="left", padx=10)

        # 统计面板
        stats_frame = ttk.LabelFrame(self.root, text="2. 验证统计", padding=10)
        stats_frame.pack(fill="x", padx=10, pady=5)

        self.total_lbl = ttk.Label(stats_frame, text="待验证：0 条")
        self.total_lbl.grid(row=0, column=0, padx=20, pady=5)

        self.verified_lbl = ttk.Label(stats_frame, text="已验证：0 条")
        self.verified_lbl.grid(row=0, column=1, padx=20, pady=5)

        self.matched_lbl = ttk.Label(stats_frame, text="✅ 匹配：0")
        self.matched_lbl.grid(row=0, column=2, padx=20, pady=5)

        self.mismatch_lbl = ttk.Label(stats_frame, text="❌ 不匹配：0")
        self.mismatch_lbl.grid(row=0, column=3, padx=20, pady=5)

        self.error_lbl = ttk.Label(stats_frame, text="🔴 错误：0")
        self.error_lbl.grid(row=0, column=4, padx=20, pady=5)

        # 进度条
        progress_frame = ttk.Frame(self.root, padding=10)
        progress_frame.pack(fill="x", padx=10, pady=5)

        self.progress = ttk.Progressbar(progress_frame, mode="determinate")
        self.progress.pack(fill="x")

        self.progress_lbl = ttk.Label(progress_frame, text="进度：0.00%")
        self.progress_lbl.pack(anchor="w")

        # 控制按钮
        control_frame = ttk.Frame(self.root, padding=10)
        control_frame.pack(fill="x", padx=10, pady=5)

        self.start_btn = ttk.Button(
            control_frame, text="▶️ 开始验证", command=self._start_verification, state="disabled"
        )
        self.start_btn.pack(side="left", padx=5)

        self.pause_btn = ttk.Button(control_frame, text="⏸️ 暂停", command=self._toggle_pause, state="disabled")
        self.pause_btn.pack(side="left", padx=5)

        self.stop_btn = ttk.Button(control_frame, text="⏹️ 停止", command=self._stop_verification, state="disabled")
        self.stop_btn.pack(side="left", padx=5)

        report_btn = ttk.Button(control_frame, text="📊 生成报告", command=self._generate_report, state="disabled")
        report_btn.pack(side="right", padx=5)

        # 日志区域
        log_frame = ttk.LabelFrame(self.root, text="3. 验证日志", padding=10)
        log_frame.pack(fill="both", expand=True, padx=10, pady=5)

        self.log_text = scrolledtext.ScrolledText(log_frame, height=15, wrap="word", state="disabled")
        self.log_text.pack(fill="both", expand=True)

        # 状态栏
        self.status_bar = ttk.Label(self.root, text="就绪", relief="sunken")
        self.status_bar.pack(fill="x", padx=10, pady=5)

    def _browse_file(self):
        """浏览选择文件"""
        filename = filedialog.askopenfilename(
            title="选择演员数据库 Excel 文件", filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")]
        )
        if filename:
            self.file_var.set(filename)

    def _load_file(self):
        """加载文件"""
        excel_path = self.file_var.get()
        if not excel_path:
            messagebox.showwarning("警告", "请先选择 Excel 文件")
            return

        if not Path(excel_path).exists():
            messagebox.showerror("错误", f"文件不存在：{excel_path}")
            return

        self._log(f"加载文件：{excel_path}")

        try:
            self.verifier = TMDBVerifier(excel_path)
            count = self.verifier.find_records_to_verify()

            self.total_lbl.config(text=f"待验证：{count} 条")
            self.start_btn.config(state="normal")
            self.report_btn.config(state="normal")

            self._log(f"✅ 加载成功，找到 {count} 条待验证记录")
            self.status_bar.config(text=f"已加载：{excel_path} (待验证：{count} 条)")

        except Exception as e:
            messagebox.showerror("错误", f"加载文件失败：{e}")
            self._log(f"❌ 加载失败：{e}")

    def _start_verification(self):
        """开始验证"""
        if not self.verifier:
            return

        self._log("\n" + "=" * 60)
        self._log("开始验证...")

        self.is_running = True
        self.start_btn.config(state="disabled")
        self.pause_btn.config(state="normal")
        self.stop_btn.config(state="normal")

        # 启动验证线程
        import threading

        thread = threading.Thread(target=self._run_verification, daemon=True)
        thread.start()

    def _run_verification(self):
        """运行验证（后台线程）"""
        total = len(self.verifier.to_verify)

        for i, (row_idx, jp_name, tmdbid) in enumerate(self.verifier.to_verify):
            if not self.is_running:
                self._log("⏹️ 验证已停止")
                break

            # 验证
            status, result = self.verifier.verify_single(row_idx, jp_name, tmdbid)

            # 更新统计
            self.verifier.stats["verified"] += 1
            if status == "matched":
                self.verifier.stats["matched"] += 1
                full_status = "✅ 匹配"
            elif status == "mismatch":
                self.verifier.stats["mismatch"] += 1
                full_status = "❌ 不匹配"
            elif status == "no_network":
                self.verifier.stats["no_network"] += 1
                full_status = "⚠️ 无网络"
            else:
                self.verifier.stats["error"] += 1
                full_status = "🔴 错误"

            self.verifier.results.append((row_idx, jp_name, tmdbid, status, result))

            # 更新 Excel
            self.verifier.update_cell(row_idx, full_status, result)

            # 定期保存
            if (i + 1) % self.save_interval == 0:
                self.verifier.save_progress(i + 1)
                self._save_excel()

            # 更新 UI
            verified = self.verifier.stats["verified"]
            self.verified_lbl.config(text=f"已验证：{verified}")
            self.matched_lbl.config(text=f"✅ 匹配：{self.verifier.stats['matched']}")
            self.mismatch_lbl.config(text=f"❌ 不匹配：{self.verifier.stats['mismatch']}")
            self.error_lbl.config(text=f"🔴 错误：{self.verifier.stats['error']}")

            progress = (i + 1) / total * 100
            self.progress["value"] = progress
            self.progress_lbl.config(text=f"进度：{progress:.2f}% ({i + 1}/{total})")

            # 日志
            self._log(f"[{i + 1}/{total}] {jp_name} -> {result}")
            self.status_bar.config(text=f"正在验证：{jp_name}")

        # 完成
        if self.is_running:
            self.verifier.save_progress(total)
            self._save_excel()
            self._log("\n✅ 验证完成！")
            self._log(
                f"总计：{total} 条 | 匹配：{self.verifier.stats['matched']} | 不匹配：{self.verifier.stats['mismatch']}"
            )
            messagebox.showinfo(
                "完成",
                f"验证完成！\n\n匹配：{self.verifier.stats['matched']}\n不匹配：{self.verifier.stats['mismatch']}\n错误：{self.verifier.stats['error']}",
            )

        self.start_btn.config(state="normal")
        self.pause_btn.config(state="disabled")
        self.stop_btn.config(state="disabled")
        self.status_bar.config(text="验证完成")

    def _toggle_pause(self):
        """暂停/继续"""
        if self.is_running:
            self.is_running = False
            self.pause_btn.config(text="▶️ 继续")
            self._log("⏸️ 验证已暂停")
        else:
            self.is_running = True
            self.pause_btn.config(text="⏸️ 暂停")
            self._log("▶️ 验证已继续")
            import threading

            thread = threading.Thread(target=self._run_verification, daemon=True)
            thread.start()

    def _stop_verification(self):
        """停止"""
        self.is_running = False
        self.start_btn.config(state="normal")
        self.pause_btn.config(state="disabled")
        self.stop_btn.config(state="disabled")
        self.pause_btn.config(text="⏸️ 暂停")
        self._log("⏹️ 验证已停止")

    def _save_excel(self):
        """保存 Excel"""
        try:
            if self.verifier and self.verifier.wb:
                self.verifier.wb.save(self.verifier.excel_path)
        except Exception as e:
            self._log(f"⚠️ 保存失败：{e}")

    def _generate_report(self):
        """生成报告"""
        if not self.verifier or not self.verifier.results:
            messagebox.showwarning("警告", "还没有验证数据，请先运行验证")
            return

        try:
            report_path = self.verifier.generate_report()
            self._log(f"📊 报告已生成：{report_path}")
            messagebox.showinfo("完成", f"报告已生成：\n{report_path}")
        except Exception as e:
            messagebox.showerror("错误", f"生成报告失败：{e}")

    def _log(self, message: str):
        """添加日志"""
        self.log_text.config(state="normal")
        timestamp = datetime.now().strftime("%H:%M:%S")
        self.log_text.insert("end", f"[{timestamp}] {message}\n")
        self.log_text.see("end")
        self.log_text.config(state="disabled")


def main():
    """主函数"""
    root = tk.Tk()

    # 设置样式
    style = ttk.Style()
    style.theme_use("clam")

    # 设置窗口图标（可选）
    try:
        root.iconbitmap(default="")
    except Exception:
        pass

    VerifierGUI(root)
    root.mainloop()


if __name__ == "__main__":
    # 检查是否直接运行
    if len(sys.argv) > 1:
        # 命令行模式：直接验证指定文件
        excel_path = sys.argv[1]
        sample_size = int(sys.argv[2]) if len(sys.argv) > 2 else 0

        verifier = TMDBVerifier(excel_path)
        count = verifier.find_records_to_verify()

        print(f"📋 找到 {count} 条待验证记录")

        if sample_size > 0:
            verifier.to_verify = verifier.to_verify[:sample_size]

        if count == 0:
            print("✅ 所有记录已验证完毕")
            sys.exit(0)

        # 运行验证
        total = len(verifier.to_verify)
        for i, (row_idx, jp_name, tmdbid) in enumerate(verifier.to_verify):
            status, result = verifier.verify_single(row_idx, jp_name, tmdbid)

            verifier.stats["verified"] += 1
            if status == "matched":
                verifier.stats["matched"] += 1
            elif status == "mismatch":
                verifier.stats["mismatch"] += 1
            elif status == "no_network":
                verifier.stats["no_network"] += 1
            else:
                verifier.stats["error"] += 1

            verifier.update_cell(row_idx, status, result)

            if (i + 1) % 200 == 0:
                verifier.save_progress(i + 1)

            print(f"[{i + 1}/{total}] {jp_name} -> {result}")

        verifier.save_progress(total)

        # 生成报告
        report_path = verifier.generate_report()
        print(f"\n✅ 验证完成！报告：{report_path}")

        # 打印不匹配记录
        mismatches = [r for r in verifier.results if r[3] == "mismatch"]
        if mismatches:
            print(f"\n❌ 发现 {len(mismatches)} 条不匹配记录:")
            for row, jp, tid, status, result in mismatches[:20]:
                print(f"  行{row}: {jp} -> TMDB ID:{tid} ({result})")
    else:
        # GUI 模式
        main()
