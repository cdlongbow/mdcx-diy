"""
演员 TMDB ID 查询模块。

通过 TMDB API 搜索日本成人演员，获取 tmdbid 及多语言翻译，统一管理演员数据库 xlsx。
"""

import asyncio
import json
import os
import time
from pathlib import Path
from typing import Any

import zhconv
from curl_cffi.requests import AsyncSession
from parsel import Selector

from ..config.manager import manager
from ..config.resources import (
    COL_HREF,
    COL_KEYWORD,
    COL_TMDB_URL,
    COL_TMDBID,
    COL_ZH_CN,
    COL_ZH_TW,
    DB_HEADERS,
    _tmdb_person_url,
    read_actor_db_xlsx,
    resources,
)
from ..models.log_buffer import LogBuffer
from ..utils import convert_half

# 演员数据库写锁：防止多个并发任务同时写 xlsx 导致文件损坏
_actor_db_write_lock = asyncio.Lock()

# TMDB 查询内存缓存与并发去重（避免同名演员重复请求）
_TMDB_QUERY_CACHE_TTL_SECONDS = 24 * 60 * 60
_TMDB_QUERY_CACHE_MAX_ENTRIES = 2000
_TMDB_QUERY_CACHE: dict[str, tuple[float, dict | None]] = {}
_TMDB_QUERY_INFLIGHT: dict[str, asyncio.Task[dict | None]] = {}
_TMDB_QUERY_STATE_LOCK = asyncio.Lock()
_TMDB_QUERY_CACHE_IO_LOCK = asyncio.Lock()


async def _read_text_file(path: Path, encoding: str = "utf-8") -> str:
    return await asyncio.to_thread(path.read_text, encoding=encoding)


# ============= 速率限制器 =============


class _TmdbRateLimiter:
    """令牌桶限流器，控制 TMDB API 请求速率。"""

    def __init__(self, rate: float = 3.5, burst: int = 10, min_rate: float = 1.0, max_rate: float = 6.0):
        self._base_rate = float(rate)
        self._rate = float(rate)
        self._min_rate = float(min_rate)
        self._max_rate = float(max_rate)
        self._tokens = float(burst)
        self._max_tokens = float(burst)
        self._last = time.monotonic()
        self._lock = asyncio.Lock()
        self._last_success = time.monotonic()
        self._consecutive_errors = 0
        self._inflight_requests = 0

    def _effective_rate(self) -> float:
        load_factor = 1.0 + (self._inflight_requests / 8)
        return max(self._min_rate, self._rate / load_factor)

    def _calc_rate(self, ok: bool) -> float:
        if not ok:
            # 遇到错误时迅速降速，保守重试
            self._consecutive_errors = min(self._consecutive_errors + 1, 8)
            return max(self._min_rate, self._rate * (0.7**self._consecutive_errors))

        self._consecutive_errors = 0
        elapsed = time.monotonic() - self._last_success
        if elapsed >= 1.5:
            # 稳定成功后慢慢恢复速率
            return min(self._max_rate, self._rate + 0.2)
        return self._rate

    def _reclaim_tokens_locked(self, now: float) -> None:
        self._tokens = min(self._max_tokens, self._tokens + (now - self._last) * self._rate)
        self._last = now

    async def acquire(self) -> None:
        async with self._lock:
            self._inflight_requests += 1
            now = time.monotonic()
            self._reclaim_tokens_locked(now)
            effective_rate = self._effective_rate()
            if self._tokens < 1:
                delay = (1 - self._tokens) / effective_rate
                await asyncio.sleep(delay)
                self._tokens = 0
            else:
                self._tokens -= 1

    async def finish(self, status_code: int) -> None:
        async with self._lock:
            self._inflight_requests = max(0, self._inflight_requests - 1)
            now = time.monotonic()
            self._reclaim_tokens_locked(now)
            if 200 <= status_code < 300:
                self._last_success = now
                self._rate = self._calc_rate(ok=True)
                return

            if status_code == 429:
                # TMDB 限流时快速降速
                self._rate = max(self._min_rate, self._rate * 0.45)
                self._consecutive_errors = min(self._consecutive_errors + 2, 8)
                return

            self._rate = self._calc_rate(ok=False)


def _tmdb_query_cache_key(actor_name: str) -> str:
    candidates = _actor_index_keys(actor_name)
    if not candidates:
        return norm_name(actor_name).upper()
    return sorted(candidates, key=len)[0]


def _tmdb_query_cache_get(name: str) -> dict | None:
    cached = _TMDB_QUERY_CACHE.get(name)
    if not cached:
        return None
    ts, value = cached
    if time.time() - ts > _TMDB_QUERY_CACHE_TTL_SECONDS:
        _TMDB_QUERY_CACHE.pop(name, None)
        return None
    return value


def _tmdb_query_cache_set(name: str, value: dict | None) -> None:
    _tmdb_query_cache_set_and_persist(name, value)


def _tmdb_query_cache_set_and_persist(name: str, value: dict | None) -> None:
    if not isinstance(name, str):
        return
    if not isinstance(value, (dict, type(None))):
        return
    _TMDB_QUERY_CACHE[name] = (time.time(), value)
    _tmdb_query_cache_sanitize()

    try:
        loop = asyncio.get_running_loop()
    except RuntimeError:
        _tmdb_query_cache_persist_sync()
    else:
        loop.create_task(_tmdb_query_cache_persist_async())


def _tmdb_query_cache_path() -> Path:
    cache_path = getattr(resources, "tmdb_query_cache_path", None)
    if isinstance(cache_path, Path):
        if cache_path == Path() or str(cache_path) in (".", "./"):
            return _tmdb_query_cache_fallback_path()
        return cache_path
    if not cache_path:
        return _tmdb_query_cache_fallback_path()
    candidate = Path(cache_path)
    if candidate == Path() or str(candidate) in (".", "./"):
        return _tmdb_query_cache_fallback_path()
    if not candidate.is_absolute():
        return _tmdb_query_cache_fallback_path()
    return candidate


def _tmdb_query_cache_fallback_path() -> Path | None:
    data_folder = getattr(manager, "data_folder", None)
    base = data_folder if isinstance(data_folder, Path) else Path(str(data_folder)) if data_folder else Path.cwd()
    return base / "userdata" / "actor_tmdb_query_cache.json"


def _tmdb_query_cache_sanitize(now: float | None = None) -> None:
    if now is None:
        now = time.time()

    entries = _TMDB_QUERY_CACHE.items()
    valid: dict[str, tuple[float, dict | None]] = {}
    for name, cached in entries:
        if not isinstance(name, str):
            continue
        if not isinstance(cached, tuple) or len(cached) != 2:
            continue
        ts, value = cached
        if not isinstance(ts, (int, float)):
            continue
        if not isinstance(value, (dict, type(None))):
            continue
        if now - float(ts) > _TMDB_QUERY_CACHE_TTL_SECONDS:
            continue
        valid[name] = (float(ts), value)

    if not valid:
        _TMDB_QUERY_CACHE.clear()
        return

    if len(valid) > _TMDB_QUERY_CACHE_MAX_ENTRIES:
        # keep latest update time
        valid = dict(sorted(valid.items(), key=lambda kv: kv[1][0], reverse=True)[:_TMDB_QUERY_CACHE_MAX_ENTRIES])

    _TMDB_QUERY_CACHE.clear()
    _TMDB_QUERY_CACHE.update(valid)


def _tmdb_query_cache_load() -> None:
    cache_path = _tmdb_query_cache_path()
    if cache_path is None or not cache_path.exists():
        return

    try:
        raw = cache_path.read_text(encoding="utf-8")
        data = json.loads(raw) if raw.strip() else {}
    except Exception:
        return

    if not isinstance(data, dict):
        return

    now = time.time()
    for name, payload in data.items():
        if not isinstance(name, str) or not isinstance(payload, dict):
            continue

        ts = payload.get("ts")
        value = payload.get("value")
        if not isinstance(ts, (int, float)):
            continue
        if value is not None and not isinstance(value, dict):
            continue

        _TMDB_QUERY_CACHE[name] = (float(ts), value)

    _tmdb_query_cache_sanitize(now=now)


async def _tmdb_query_cache_persist_async() -> None:
    async with _TMDB_QUERY_CACHE_IO_LOCK:
        _tmdb_query_cache_persist_sync()


def _tmdb_query_cache_persist_sync() -> None:
    cache_path = _tmdb_query_cache_path()
    if not cache_path:
        return

    now = time.time()
    payload = {}
    _tmdb_query_cache_sanitize()
    for name, (ts, value) in _TMDB_QUERY_CACHE.items():
        if now - ts > _TMDB_QUERY_CACHE_TTL_SECONDS:
            continue
        payload[name] = {"ts": ts, "value": value}

    cache_path.parent.mkdir(parents=True, exist_ok=True)
    tmp_path = cache_path.with_suffix(cache_path.suffix + ".tmp")
    tmp_path.write_text(json.dumps(payload, ensure_ascii=False), encoding="utf-8")
    os.replace(tmp_path, cache_path)


_tmdb_query_cache_load()


def flush_tmdb_query_cache() -> None:
    """Flush pending TMDB query cache data to disk immediately."""
    _tmdb_query_cache_persist_sync()


_tmdb_rate_limiter = _TmdbRateLimiter()


def _tmdb_debug_enabled() -> bool:
    return bool(getattr(manager.config, "show_data_log", False))


def _tmdb_log_line(message: str) -> None:
    """Write one TMDB log entry as a separate line."""
    if message.startswith("\n"):
        LogBuffer.log().write(message)
    else:
        LogBuffer.log().write(f"\n{message}")


# ============= HTTP 适配器 =============


class _TmdbResponse:
    """Unified response wrapper for TMDB API calls."""

    def __init__(self, status_code: int, text: str):
        self.status_code = status_code
        self._text = text

    def json(self) -> dict[str, Any]:
        return json.loads(self._text)


async def _tmdb_request(client: Any, method: str, url: str, **kwargs) -> _TmdbResponse | None:
    """
    Unified HTTP request for both curl_cffi AsyncWebClient and aiohttp ClientSession.
    Returns _TmdbResponse or None on failure.
    """
    await _tmdb_rate_limiter.acquire()
    params = kwargs.get("params")
    follow_redirects = kwargs.get("follow_redirects", True)
    status_code = 0
    try:
        if hasattr(client, "request"):
            resp, err = await client.request(method, url, params=params)
            if resp is None:
                return None
            status_code = int(resp.status_code)
            return _TmdbResponse(status_code, resp.text)
        elif hasattr(client, method.lower()):
            send = getattr(client, method.lower())
            resp = await send(url, params=params, allow_redirects=follow_redirects)
            status_code = int(resp.status)
            return _TmdbResponse(status_code, await resp.text())
    finally:
        await _tmdb_rate_limiter.finish(status_code)
    return None


def is_japan_place(place: str) -> bool:
    if not place:
        return False
    p = place.lower()
    japan_keywords = [
        "japan",
        "日本",
        "北海道",
        "青森",
        "岩手",
        "宫城",
        "秋田",
        "山形",
        "福岛",
        "茨城",
        "栃木",
        "群马",
        "埼玉",
        "千叶",
        "东京",
        "神奈川",
        "新潟",
        "富山",
        "石川",
        "福井",
        "山梨",
        "长野",
        "岐阜",
        "静冈",
        "爱知",
        "三重",
        "滋贺",
        "京都",
        "大阪",
        "兵库",
        "奈良",
        "和歌山",
        "鸟取",
        "岛根",
        "冈山",
        "广岛",
        "山口",
        "德岛",
        "香川",
        "爱媛",
        "高知",
        "福冈",
        "佐贺",
        "长崎",
        "熊本",
        "大分",
        "宫崎",
        "鹿儿岛",
        "冲绳",
        "横滨",
        "札幌",
        "神户",
        "川崎",
        "仙台",
        "町田",
        "八王子",
        "品川",
        "涩谷",
        "新宿",
        "池袋",
        "浅草",
        "秋叶原",
        "原宿",
        "六本木",
        "银座",
        "上野",
        "下北泽",
        "吉祥寺",
        "武藏野",
        "国分寺",
        "立川",
        "日野",
        "多摩",
        "青梅",
        # 常见日语地名后缀（仅当长度>=3时匹配，避免单字误匹配中文）
        "都",
        "道",
        "府",
        "县",
        "市",
        "町",
        "村",
        "区",
    ]
    # 排除明显非日本籍（中国地址中 "都" 常见于 "广东/成都"）
    if "中国" in p or "china" in p:
        return False
    for kw in japan_keywords:
        if len(kw) >= 2 and kw in p:
            return True
        if len(kw) == 1 and len(p) >= 3 and p.endswith(kw):
            return True
    return False


# ============= 演员数据库 xlsx 管理 =============

# ============= 演员名称规范化和变体扩展 =============

_ACTOR_NAME_VARIANT_MAP = str.maketrans(
    {
        "亜": "亚",
        "亞": "亚",
        "条": "条",
        "條": "条",
        "瀬": "濑",
        "濑": "濑",
        "沢": "泽",
        "澤": "泽",
        "桜": "樱",
        "櫻": "樱",
        "髙": "高",
        "﨑": "崎",
    }
)


def norm_name(name: str) -> str:
    return convert_half(name or "").translate(_ACTOR_NAME_VARIANT_MAP)


def _expand_name_variants(name: str) -> set[str]:
    normalized = norm_name(name)
    if not normalized:
        return set()

    variants = {normalized}

    if normalized.endswith("こ"):
        variants.add(normalized[:-1] + "子")
    if normalized.endswith("子"):
        variants.add(normalized[:-1] + "こ")

    return variants


def _get_db_path() -> Path:
    return manager.data_folder / "userdata" / "actor_database.xlsx"


def _format_db_worksheet(ws) -> None:
    """格式化演员数据库工作表：固定表头、自动筛选、列宽、边框、超链接、表头样式。"""
    try:
        import openpyxl
        from openpyxl.utils import get_column_letter

        # 固定表头
        ws.freeze_panes = "B2"

        # 自动筛选
        last_col = get_column_letter(len(DB_HEADERS))
        max_row = ws.max_row if ws.max_row else 1
        ws.auto_filter.ref = f"A1:{last_col}{max_row}"

        # 表头样式
        header_fill = openpyxl.styles.PatternFill("solid", fgColor="F2F2F2")
        header_font = openpyxl.styles.Font(bold=True, size=11)
        header_align = openpyxl.styles.Alignment(horizontal="center", vertical="center", wrap_text=True)
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_align

        # 表格边框
        thin = openpyxl.styles.Side(style="thin", color="D0D0D0")
        border = openpyxl.styles.Border(left=thin, right=thin, top=thin, bottom=thin)
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=len(DB_HEADERS)):
            for cell in row:
                cell.border = border

        # 链接列和 tmdb url 列超链接：每次保存都校验并修复
        for row in ws.iter_rows(min_row=2, values_only=False):
            href_cell = row[COL_HREF]
            tmdb_cell = row[COL_TMDB_URL]
            href_val = str(href_cell.value or "").strip()
            if href_val and href_val.startswith("http"):
                existing_target = href_cell.hyperlink.target if href_cell.hyperlink else None
                if existing_target != href_val:
                    href_cell.style = "Hyperlink"
                    href_cell.hyperlink = href_val
            tmdbid_val = str(row[COL_TMDBID].value or "").strip()
            if tmdbid_val.isdigit():
                expected_tmdb_val = _tmdb_person_url(tmdbid_val)
                if str(tmdb_cell.value or "").strip() != expected_tmdb_val:
                    tmdb_cell.value = expected_tmdb_val
                tmdb_val = expected_tmdb_val
            else:
                tmdb_val = str(tmdb_cell.value or "").strip()
            if tmdb_val and tmdb_val.startswith("http"):
                existing_target = tmdb_cell.hyperlink.target if tmdb_cell.hyperlink else None
                if existing_target != tmdb_val:
                    tmdb_cell.style = "Hyperlink"
                    tmdb_cell.hyperlink = tmdb_val

        # 超链接处理会覆盖边框，重新设置
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=len(DB_HEADERS)):
            for cell in row:
                cell.border = border

        # 自动列宽
        caps = {1: 25, 2: 15, 3: 15, 4: 60, 5: 50, 6: 12, 7: 42}
        col_max = [0] * (len(DB_HEADERS) + 1)
        for row in ws.iter_rows(min_row=2, values_only=True):
            for ci, cell in enumerate(row, 1):
                if cell is None or ci > len(DB_HEADERS):
                    continue
                s = str(cell)
                width = sum(2 if "\u3040" <= c <= "\u30ff" or "\u4e00" <= c <= "\u9fff" else 1 for c in s)
                col_max[ci] = max(col_max[ci], width)
        for ci in range(1, len(DB_HEADERS) + 1):
            letter = get_column_letter(ci)
            ws.column_dimensions[letter].width = min(col_max[ci] + 2, caps.get(ci, 80))
    except Exception as e:
        LogBuffer.log().write(f"  ⚠️ [演员数据库] 工作表格式化失败: {e}")


def _norm_name_set(names: list[str]) -> set[str]:
    """生成一组规范化后的名字（全半角统一、零宽字符清理、大小写统一）"""
    result = set()
    for n in names:
        if n:
            result.update(_expand_name_variants(n))
    return result


def _full_to_half(s: str) -> str:
    result = []
    for ch in s:
        cp = ord(ch)
        if cp == 0x3000:
            result.append(" ")
        elif 0xFF01 <= cp <= 0xFF5E:
            result.append(chr(cp - 0xFEE0))
        else:
            result.append(ch)
    return "".join(result)


def _actor_index_keys(name: str) -> set[str]:
    if not name:
        return set()
    variants = _expand_name_variants(name)
    result = set(variants)
    result.add(_full_to_half(norm_name(name)).upper())
    for variant in variants:
        result.add(_full_to_half(variant).upper())
    return {key for key in result if key}


def _merge_keyword_values(*groups: list[str] | tuple[str, ...] | set[str] | str) -> str:
    merged: list[str] = []
    seen: set[str] = set()

    for group in groups:
        values = [group] if isinstance(group, str) else list(group)
        for value in values:
            normalized = str(value or "").strip()
            if not normalized or normalized in seen:
                continue
            seen.add(normalized)
            merged.append(normalized)

    return ",".join(merged)


def _build_actor_db_reverse_index(actor_db: dict[str, dict]) -> dict[str, str]:
    index: dict[str, str] = {}
    for jp, data in actor_db.items():
        candidates = [jp]
        if data.get("zh_cn"):
            candidates.append(data["zh_cn"])
        if data.get("zh_tw"):
            candidates.append(data["zh_tw"])
        if data.get("keyword"):
            candidates.extend(k.strip() for k in data["keyword"].split(",") if k.strip())

        for candidate in candidates:
            for key in _actor_index_keys(candidate):
                index.setdefault(key, jp)
    return index


async def load_actor_db() -> dict[str, dict]:
    """
    加载演员数据库 xlsx。
    返回: {jp_name: {"zh_cn": ..., "zh_tw": ..., "keyword": "...", "href": "...", "tmdbid": int|None, "tmdb_url": "..."}}
    """
    db_path = _get_db_path()
    if not db_path.exists():
        return {}
    try:
        return read_actor_db_xlsx(db_path)
    except ImportError:
        LogBuffer.log().write("  ⚠️ [演员数据库] 缺少 openpyxl，无法读取 actor_database.xlsx")
    except Exception as e:
        LogBuffer.log().write(f"  ⚠️ [演员数据库] 读取失败: {e}")
    return {}


async def update_actor_db_row(
    jp: str,
    zh_cn: str = "",
    zh_tw: str = "",
    keyword: str = "",
    href: str = "",
    tmdbid: int | None = None,
    append_keyword: bool = False,
    overwrite_names: bool = False,
) -> str:
    """
    更新或添加演员数据库行。

    默认已有值不覆盖，仅填空白；keyword 在 append_keyword=True 时追加去重。
    当 overwrite_names=True 时，zh_cn/zh_tw 允许用新值覆盖已有值（用于已有 tmdbid 演员翻译补全）。
    """
    db_path = _get_db_path()
    db_path.parent.mkdir(parents=True, exist_ok=True)
    async with _actor_db_write_lock:
        try:
            import openpyxl

            write_status = "unchanged"

            if db_path.exists():
                wb = openpyxl.load_workbook(db_path)
            else:
                wb = openpyxl.Workbook()
                ws = wb.active
                ws.title = "演员数据库"
                for col, header in enumerate(DB_HEADERS, 1):
                    cell = ws.cell(row=1, column=col, value=header)
                    cell.font = openpyxl.styles.Font(bold=True)
                    cell.fill = openpyxl.styles.PatternFill("solid", fgColor="C0C0C0")
                    cell.alignment = openpyxl.styles.Alignment(horizontal="center")

            ws = wb.active
            if ws.title != "演员数据库":
                ws.title = "演员数据库"

            existing_row = None
            for row_idx, row in enumerate(ws.iter_rows(min_col=1, max_col=7, values_only=True), start=1):
                if row_idx == 1:
                    continue
                if str(row[0] or "").strip() == jp:
                    existing_row = row_idx
                    break

            if existing_row:
                existing_zh_cn = ws.cell(row=existing_row, column=COL_ZH_CN + 1).value
                existing_zh_tw = ws.cell(row=existing_row, column=COL_ZH_TW + 1).value
                if overwrite_names:
                    if zh_cn and str(existing_zh_cn or "").strip() != zh_cn:
                        ws.cell(row=existing_row, column=COL_ZH_CN + 1, value=zh_cn)
                        write_status = "updated_zh_cn"
                    if zh_tw and str(existing_zh_tw or "").strip() != zh_tw:
                        ws.cell(row=existing_row, column=COL_ZH_TW + 1, value=zh_tw)
                        write_status = "updated_zh_cn_zh_tw" if write_status == "updated_zh_cn" else "updated_zh_tw"
                else:
                    if not existing_zh_cn and zh_cn:
                        ws.cell(row=existing_row, column=COL_ZH_CN + 1, value=zh_cn)
                    if not existing_zh_tw and zh_tw:
                        ws.cell(row=existing_row, column=COL_ZH_TW + 1, value=zh_tw)

                if keyword:
                    existing_kw = str(ws.cell(row=existing_row, column=COL_KEYWORD + 1).value or "").strip()
                    if append_keyword:
                        new_kws = {k.strip() for k in keyword.split(",") if k.strip()}
                        existing_kws_set = set()
                        if existing_kw:
                            existing_kws_set = {k.strip() for k in existing_kw.split(",") if k.strip()}
                        merged = existing_kws_set | new_kws
                        ws.cell(row=existing_row, column=COL_KEYWORD + 1, value=",".join(sorted(merged)))
                    elif not existing_kw:
                        ws.cell(row=existing_row, column=COL_KEYWORD + 1, value=keyword)

                if not ws.cell(row=existing_row, column=COL_HREF + 1).value and href:
                    ws.cell(row=existing_row, column=COL_HREF + 1, value=href)
                if tmdbid is not None and not ws.cell(row=existing_row, column=COL_TMDBID + 1).value:
                    ws.cell(row=existing_row, column=COL_TMDBID + 1, value=tmdbid)
                    write_status = "inserted_tmdbid"
                    tmdb_url = _tmdb_person_url(tmdbid)
                    ws.cell(row=existing_row, column=COL_TMDB_URL + 1).value = None
                    ws.cell(row=existing_row, column=COL_TMDB_URL + 1, value=tmdb_url)
                    ws.cell(row=existing_row, column=COL_TMDB_URL + 1).hyperlink = tmdb_url
                elif tmdbid is not None and write_status == "unchanged":
                    write_status = "kept_existing_tmdbid"
            else:
                ws.append([jp, zh_cn, zh_tw, keyword, href, tmdbid or "", ""])
                write_status = "inserted_new_row"
                last_row = ws.max_row
                if tmdbid:
                    tmdb_url = _tmdb_person_url(tmdbid)
                    ws.cell(row=last_row, column=COL_TMDB_URL + 1).value = None
                    ws.cell(row=last_row, column=COL_TMDB_URL + 1, value=tmdb_url)
                    ws.cell(row=last_row, column=COL_TMDB_URL + 1).hyperlink = tmdb_url

            _format_db_worksheet(ws)

            wb.save(db_path)
            wb.close()
            resources.reload_actor_db()
            return write_status
        except ImportError:
            return "missing_openpyxl"
        except PermissionError:
            return "file_locked"
        except Exception as e:
            return f"write_failed:{e}"


def search_actor_db_reverse(query_name: str) -> dict | None:
    """
    反向搜索演员数据库：用任意语言的演员名搜索。返回匹配的整行数据，或 None。
    """
    actor_db = resources.actor_db
    if not actor_db:
        return None

    reverse_index = getattr(resources, "actor_db_reverse_index", None)
    if reverse_index is None:
        reverse_index = _build_actor_db_reverse_index(actor_db)
        resources.actor_db_reverse_index = reverse_index

    for key in _actor_index_keys(query_name):
        jp = reverse_index.get(key)
        if jp and jp in actor_db:
            data_copy = dict(actor_db[jp])
            data_copy["jp"] = jp
            return data_copy

    return None


# ============= XML → xlsx 迁移 =============


async def migrate_xml_to_xlsx() -> bool:
    """
    将 mapping_actor.xml + actor_tmdbid.xlsx 迁移到 actor_database.xlsx。
    迁移成功后返回 True。
    """
    db_path = _get_db_path()
    if db_path.exists():
        return True

    xml_path = manager.data_folder / "userdata" / "mapping_actor.xml"
    old_tmdb_path = manager.data_folder / "userdata" / "actor_tmdbid.xlsx"

    migrated = False
    try:
        if xml_path.exists():
            parser = None
            try:
                from lxml import etree

                parser = etree.HTMLParser(encoding="utf-8")
            except ImportError:
                pass
            if parser:
                content = await _read_text_file(xml_path)
                xml_data = etree.HTML(content.encode("utf-8"), parser)
                actor_objects = xml_data.xpath("//a")

                old_tmdb_cache = {}
                if old_tmdb_path.exists():
                    try:
                        import openpyxl

                        wb = openpyxl.load_workbook(old_tmdb_path, read_only=True, data_only=True)
                        ws = wb.active
                        for row_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
                            if row_idx == 1:
                                continue
                            if len(row) >= 2:
                                name = str(row[0] or "").strip()
                                tid = str(row[1] or "").strip()
                                if name and tid.isdigit():
                                    old_tmdb_cache[name] = int(tid)
                        wb.close()
                    except Exception:
                        LogBuffer.log().write("  ⚠️ [演员数据库] 旧版 tmdbid xlsx 读取失败，跳过迁移")

                import openpyxl

                wb = openpyxl.Workbook()
                ws = wb.active
                ws.title = "演员数据库"
                for col, header in enumerate(DB_HEADERS, 1):
                    cell = ws.cell(row=1, column=col, value=header)
                    cell.font = openpyxl.styles.Font(bold=True)
                    cell.fill = openpyxl.styles.PatternFill("solid", fgColor="C0C0C0")
                    cell.alignment = openpyxl.styles.Alignment(horizontal="center")

                for ob in actor_objects:
                    jp = ob.get("jp") or ""
                    zh_cn = ob.get("zh_cn") or ""
                    zh_tw = ob.get("zh_tw") or ""
                    keyword = ob.get("keyword") or ""
                    href = ob.get("href") or ""

                    if not jp and keyword:
                        first_kw = keyword.strip(",").split(",")[0]
                        jp = first_kw

                    if not jp:
                        continue

                    tmdbid = old_tmdb_cache.get(jp)

                    ws.append([jp, zh_cn, zh_tw, keyword, href, tmdbid or "", ""])
                    if tmdbid:
                        last_row = ws.max_row
                        tmdb_url = _tmdb_person_url(tmdbid)
                        ws.cell(row=last_row, column=COL_TMDB_URL + 1).value = None
                        ws.cell(row=last_row, column=COL_TMDB_URL + 1, value=tmdb_url)
                        ws.cell(row=last_row, column=COL_TMDB_URL + 1).hyperlink = tmdb_url

                _format_db_worksheet(ws)
                wb.save(db_path)
                wb.close()
                migrated = True
                LogBuffer.log().write(f"  ℹ️ [演员数据库] 已从 XML 迁移 {len(actor_objects)} 条记录")
    except Exception as e:
        LogBuffer.log().write(f"  ⚠️ [演员数据库] 迁移失败: {e}")

    return migrated


# ============= TMDB API 查询 =============


async def fetch_actor_tmdb_ids(actors: list[str], client: Any) -> dict[str, int]:
    if not actors:
        return {}

    tmdb_api_base = manager.config.tmdb_api_base.strip()
    tmdb_api_key = manager.config.tmdb_api_key.strip()

    if not tmdb_api_key:
        return {}

    if not tmdb_api_base:
        tmdb_api_base = "api.tmdb.org"

    protocol = "https://"
    if tmdb_api_base.startswith("http://"):
        protocol = "http://"
        tmdb_api_base = tmdb_api_base[7:]
    elif tmdb_api_base.startswith("https://"):
        protocol = "https://"
        tmdb_api_base = tmdb_api_base[8:]

    base_url = f"{protocol}{tmdb_api_base}"

    actor_db = resources.actor_db or {}

    if resources.actor_db is None:
        _tmdb_log_line("  ⚠️ [TMDB] actor_db 为 None —— 尝试重新加载演员数据库...")
        resources.reload_actor_db()
        actor_db = resources.actor_db or {}
        if resources.actor_db is not None:
            _tmdb_log_line(f"  ✅ [TMDB] actor_db 已加载 {len(actor_db)} 条记录")
        else:
            _tmdb_log_line("  ⚠️ [TMDB] actor_db 重新加载后仍为 None，将走 TMDB API 搜索")
    elif len(actor_db) == 0:
        _tmdb_log_line("  ⚠️ [TMDB] actor_db 为空 (0 条记录) —— 文件可能为空或格式不匹配")
    else:
        _tmdb_log_line(f"  ℹ️ [TMDB] actor_db 已加载 {len(actor_db)} 条记录")

    result: dict[str, int] = {}
    need_query: list[tuple[str, str]] = []
    need_translate: list[tuple[str, str, int]] = []

    for actor in actors:
        if not actor or not actor.strip():
            continue
        actor_stripped = actor.strip()

        in_actor_db = actor_stripped in actor_db
        if not in_actor_db:
            _tmdb_log_line(f"  🔍 [TMDB] '{actor_stripped}' 不在 actor_db 中 (actor_db size={len(actor_db)})")

        row = None
        if actor_stripped in actor_db and actor_db[actor_stripped].get("tmdbid"):
            result[actor_stripped] = actor_db[actor_stripped]["tmdbid"]
            row_data = actor_db[actor_stripped]
            _tmdb_log_line(
                f"  ✅ [TMDB] '{actor_stripped}' 从 actor_db 直接匹配, tmdbid={actor_db[actor_stripped]['tmdbid']}"
            )
            if not row_data.get("zh_cn") or not row_data.get("zh_tw"):
                need_translate.append((actor_stripped, actor_stripped, row_data["tmdbid"]))
            continue

        row = search_actor_db_reverse(actor_stripped)
        if row and row.get("tmdbid"):
            result[actor_stripped] = row["tmdbid"]
            _tmdb_log_line(f"  ✅ [TMDB] '{actor_stripped}' 从反查匹配, tmdbid={row['tmdbid']}")
            _tmdb_log_line(f" ℹ️ [TMDB] {actor_stripped} -> tmdbid={row['tmdbid']} (xlsx反查缓存)")
            if not row.get("zh_cn") or not row.get("zh_tw"):
                need_translate.append((actor_stripped, row.get("jp", actor_stripped), row["tmdbid"]))
            continue

        _tmdb_log_line(f"  ⚠️ [TMDB] '{actor_stripped}' 未匹配，将进入 TMDB API 搜索")
        need_query.append((actor_stripped, row.get("jp") if row and row.get("jp") else actor_stripped))

    async def _query_or_cached(actor_name: str, query_name: str) -> tuple[str, dict | None]:
        return actor_name, await query_single_actor_cached(query_name, base_url, tmdb_api_key, client)

    async def _translate_and_update(actor_name: str, jp_name: str, tid: int) -> None:
        try:
            translations = await _fetch_person_translations(tid, base_url, tmdb_api_key, client)
            zh_cn = _normalize_translation(translations.get("zh_cn", ""))
            zh_tw = _normalize_translation(translations.get("zh_tw", ""))
            if not zh_cn and zh_tw:
                zh_cn = zhconv.convert(zh_tw, "zh-cn")
            if not zh_tw and zh_cn:
                zh_tw = zhconv.convert(zh_cn, "zh-hant")
            if zh_cn or zh_tw:
                write_status = await update_actor_db_row(
                    jp=jp_name, zh_cn=zh_cn, zh_tw=zh_tw, tmdbid=tid, overwrite_names=True
                )
                _tmdb_log_line(
                    f" 🔄 [TMDB] {actor_name} 翻译补全: zh_cn={zh_cn or '-'} zh_tw={zh_tw or '-'} ({write_status})"
                )
        except Exception as e:
            _tmdb_log_line(f" ⚠️ [TMDB] {actor_name} 翻译补全失败: {e}")

    if need_translate:
        _tmdb_log_line(f" 🔄 [TMDB] 补全 {len(need_translate)} 个已有 tmdbid 演员的翻译")
        trans_semaphore = asyncio.Semaphore(3)

        async def _limited_translate(actor_name: str, jp_name: str, tid: int) -> None:
            async with trans_semaphore:
                await _translate_and_update(actor_name, jp_name, tid)

        trans_tasks = [asyncio.create_task(_limited_translate(a, j, t)) for a, j, t in need_translate]
        await asyncio.gather(*trans_tasks)

    if not need_query:
        _tmdb_log_line("  ℹ️ [TMDB] 所有演员已在 actor_db 中匹配，无需 API 查询")
        return result

    _tmdb_log_line(f" 🎬 [TMDB] 开始查询 {len(need_query)} 个演员的 TMDB ID: {[a[0] for a in need_query]}")

    async def _query_and_update(actor_name: str, query_name: str) -> None:
        try:
            _, query_result = await _query_or_cached(actor_name, query_name)
            if query_result:
                tmdbid = query_result["pid"]
                result[actor_name] = tmdbid

                translations = query_result.get("translations", {})
                zh_cn = _normalize_translation(translations.get("zh_cn", ""))
                zh_tw = _normalize_translation(translations.get("zh_tw", ""))
                if not zh_cn and zh_tw:
                    zh_cn = zhconv.convert(zh_tw, "zh-cn")
                if not zh_tw and zh_cn:
                    zh_tw = zhconv.convert(zh_cn, "zh-hant")
                aka = query_result.get("also_known_as", [])
                keyword_str = _merge_keyword_values(
                    query_result.get("name", ""),
                    query_result.get("original_name", ""),
                    aka,
                )

                # 优先保留站点原始演员名作为 jp 主名，TMDB 主名/别名只作为反查关键字。
                jp_name = query_name
                write_status = await update_actor_db_row(
                    jp=jp_name,
                    zh_cn=zh_cn,
                    zh_tw=zh_tw,
                    keyword=keyword_str,
                    tmdbid=tmdbid,
                    append_keyword=True,
                )

                _tmdb_log_line(f"  ✅ [TMDB] {actor_name} -> tmdbid={tmdbid}{f' (中文:{zh_cn})' if zh_cn else ''}")
                if write_status == "inserted_tmdbid":
                    _tmdb_log_line(f"  ✅ [演员数据库] 已写入 {actor_name} -> tmdbid={tmdbid}")
                elif write_status == "inserted_new_row":
                    _tmdb_log_line(f"  ✅ [演员数据库] 已新增 {actor_name}，并写入 tmdbid={tmdbid}")
                elif write_status == "kept_existing_tmdbid":
                    _tmdb_log_line(f"  ℹ️ [演员数据库] {actor_name} 已存在 tmdbid，保留原值")
                elif write_status == "missing_openpyxl":
                    _tmdb_log_line(f"  ⚠️ [演员数据库] 缺少 openpyxl，未写入 {actor_name} 的 tmdbid")
                elif write_status == "file_locked":
                    _tmdb_log_line(f"  ⚠️ [演员数据库] 文件被占用，未写入 {actor_name} 的 tmdbid")
                elif write_status.startswith("write_failed:"):
                    _tmdb_log_line(
                        f"  ⚠️ [演员数据库] 写入失败，未保存 {actor_name} 的 tmdbid: {write_status.split(':', 1)[1]}"
                    )
            else:
                _tmdb_log_line(f"  ❌ [TMDB] {actor_name} (搜索名:{query_name}) 未找到匹配的 TMDB 演员")
        except Exception as e:
            _tmdb_log_line(f"  ❌ [TMDB] {actor_name} 查询失败: {e}")

    semaphore = asyncio.Semaphore(3)

    async def _limited_query(actor_name: str, query_name: str) -> None:
        async with semaphore:
            await _query_and_update(actor_name, query_name)

    tasks = [asyncio.create_task(_limited_query(actor_name, query_name)) for actor_name, query_name in need_query]
    await asyncio.gather(*tasks)

    old_db = dict(resources.actor_db) if resources.actor_db else {}

    cached_before = len([a for a in actors if a.strip() in old_db and old_db[a.strip()].get("tmdbid")])
    _tmdb_log_line(
        f" 🎬 [TMDB] 查询完成: 缓存命中 {len(actors) - len(need_query)} 个, "
        f"本次匹配 {len(result) - cached_before} 个, 共 {len(result)} 个"
    )

    # 统一补链接：有 tmdbid 但 xlsx 链接列为空的演员
    actor_db = resources.actor_db or {}
    missing_link = [name for name in result if not (actor_db.get(name, {}).get("href") or "")]
    if missing_link:
        _tmdb_log_line(f" 🔗 [LibreDMM] 补全 {len(missing_link)} 个演员的链接")
        link_semaphore = asyncio.Semaphore(3)

        async def _fetch_and_update(actor_name: str) -> None:
            async with link_semaphore:
                try:
                    row = search_actor_db_reverse(actor_name)
                    jp_key = row.get("jp", actor_name) if row else actor_name
                    href = await fetch_libredmm_link(jp_key)
                    if href:
                        await update_actor_db_row(jp=jp_key, href=href)
                        _tmdb_log_line(f"  ✅ [LibreDMM] {actor_name} -> {href}")
                except Exception as e:
                    _tmdb_log_line(f"  ⚠️ [LibreDMM] {actor_name} 链接补全失败: {e}")

        tasks = [asyncio.create_task(_fetch_and_update(name)) for name in missing_link]
        await asyncio.gather(*tasks)

    return result


async def query_single_actor_cached(actor_name: str, base_url: str, api_key: str, client: Any) -> dict | None:
    """Query a single actor with in-memory dedupe + cache."""
    key = _tmdb_query_cache_key(actor_name)
    cached = _tmdb_query_cache_get(key)
    if cached is not None:
        return cached

    async with _TMDB_QUERY_STATE_LOCK:
        cached = _tmdb_query_cache_get(key)
        if cached is not None:
            return cached

        running = _TMDB_QUERY_INFLIGHT.get(key)
        if running is not None:
            return await running

        task = asyncio.create_task(_query_single_actor(actor_name, base_url, api_key, client))
        _TMDB_QUERY_INFLIGHT[key] = task

    try:
        result_data = await task
        _tmdb_query_cache_set(key, result_data)
        return result_data
    except Exception:
        _tmdb_query_cache_set(key, None)
        return None
    finally:
        _TMDB_QUERY_INFLIGHT.pop(key, None)


async def _query_single_actor(actor_name: str, base_url: str, api_key: str, client: Any) -> dict | None:
    """
    查询单个演员的 TMDB 信息。返回匹配结果或 None。
    """
    search_url = f"{base_url}/3/search/person"
    target_variants = _expand_name_variants(actor_name)

    resp = await _tmdb_request(
        client,
        "GET",
        search_url,
        params={
            "api_key": api_key,
            "query": actor_name,
            "include_adult": "true",
            "language": "zh-CN",
            "page": 1,
        },
    )

    if resp is None or resp.status_code != 200:
        return None

    try:
        data = resp.json()
    except (json.JSONDecodeError, ValueError):
        return None
    results = data.get("results", [])

    if _tmdb_debug_enabled():
        _tmdb_log_line(
            f"  🔎 [TMDB] 搜索演员「{actor_name}」返回 {len(results)} 个候选，目标规范名={sorted(target_variants)}"
        )
    else:
        _tmdb_log_line(f"  🔎 [TMDB] 搜索演员「{actor_name}」返回 {len(results)} 个候选")

    if not results:
        return None

    candidates: list[dict] = []

    for item in results[:5]:
        pid = item.get("id")
        if not pid:
            continue

        gender = item.get("gender")
        if gender not in (0, 1):
            continue

        detail_url = f"{base_url}/3/person/{pid}"
        detail_resp = await _tmdb_request(client, "GET", detail_url, params={"api_key": api_key, "language": "zh-CN"})

        if detail_resp is None or detail_resp.status_code != 200:
            continue

        try:
            detail = detail_resp.json()
        except (json.JSONDecodeError, ValueError):
            continue
        place = detail.get("place_of_birth", "")

        all_names = set()
        for x in [item.get("name"), item.get("original_name"), detail.get("name")]:
            if x:
                all_names.add(str(x).strip())
        aka_list = detail.get("also_known_as", [])
        for a in aka_list:
            if a:
                all_names.add(str(a).strip())

        all_norm = _norm_name_set(list(all_names))
        is_match = bool(target_variants & all_norm)

        _pop = item.get("popularity", 0) or 0
        popularity = float(_pop) if isinstance(_pop, (int, float, str)) else 0.0
        known_for_count = len(detail.get("known_for", [])) if "known_for" in detail else 0
        place_has_japan = is_japan_place(place)

        candidates.append(
            {
                "pid": pid,
                "is_match": is_match,
                "adult": bool(item.get("adult")),
                "popularity": popularity,
                "known_for_count": known_for_count,
                "place_has_japan": place_has_japan,
                "name": detail.get("name", ""),
                "original_name": item.get("original_name", ""),
                "raw_names": sorted(all_names),
                "all_norm": sorted(all_norm),
                "translations": {},
                "also_known_as": [a for a in aka_list if a],
            }
        )

        if _tmdb_debug_enabled():
            status = "✅" if is_match else "ℹ️"
            _tmdb_log_line(
                f"    {status} [TMDB] pid={pid} name={detail.get('name', '')} "
                f"place={place or '-'} aliases={sorted(all_names)} norm={sorted(all_norm)}"
            )

    matched = [c for c in candidates if c["is_match"]]

    if not matched:
        reason = "无候选通过名字匹配" if candidates else "候选列表为空 (搜索失败/性别过滤/detail请求失败)"
        _tmdb_log_line(f"  ⚠️ [TMDB] 演员「{actor_name}」{reason}，候选数={len(candidates)}")
        return None

    matched.sort(
        key=lambda x: (x["place_has_japan"], x["adult"], x["popularity"], x["known_for_count"]),
        reverse=True,
    )

    if len(matched) > 1:
        _tmdb_log_line(
            f"  ⚠️ [TMDB] 演员「{actor_name}」匹配到 {len(matched)} 个结果，"
            f"已自动选择 tmdbid={matched[0]['pid']} "
            f"(popularity={matched[0]['popularity']:.2f})"
        )

    selected = matched[0]
    translations = await _fetch_person_translations(selected["pid"], base_url, api_key, client)
    if translations:
        selected["translations"] = translations

    return selected


async def _fetch_person_translations(pid: int, base_url: str, api_key: str, client: Any) -> dict:
    """
    从 TMDB translations API 获取演员的多语言翻译。
    """
    result = {"zh_cn": "", "zh_tw": ""}
    try:
        trans_url = f"{base_url}/3/person/{pid}/translations"
        trans_resp = await _tmdb_request(client, "GET", trans_url, params={"api_key": api_key})
        if trans_resp is None or trans_resp.status_code != 200:
            return result

        try:
            trans_data = trans_resp.json()
        except (json.JSONDecodeError, ValueError):
            return result
        for trans in trans_data.get("translations", []):
            iso = trans.get("iso_639_1", "")
            region = trans.get("iso_3166_1", "")
            en_name = trans.get("english_name", "")
            data = trans.get("data", {})
            native_name = data.get("name", "")

            if iso == "zh" and region == "CN":
                result["zh_cn"] = _normalize_translation(data.get("name") or en_name or "")
            elif iso == "zh" and region == "TW":
                result["zh_tw"] = _normalize_translation(data.get("name") or en_name or "")
            elif iso == "zh":
                name_to_use = _normalize_translation(native_name or en_name)
                if name_to_use:
                    if not result["zh_cn"]:
                        result["zh_cn"] = name_to_use
                    if not result["zh_tw"]:
                        result["zh_tw"] = name_to_use
    except Exception:
        pass

    return result


# ============= LibreDMM 链接查询 =============


async def fetch_libredmm_link(actor_name: str) -> str:
    """搜索 LibreDMM 获取演员页面链接，未找到返回空字符串。"""
    search_url = "https://www.libredmm.com/actresses"
    params = {"order": "New", "fuzzy": actor_name, "commit": "Filter by name"}
    headers = {
        "Accept-Language": "zh-CN,zh;q=0.9",
        "Referer": "https://www.libredmm.com/actresses",
    }
    try:
        async with AsyncSession(impersonate="safari15_5") as s:
            resp = await s.get(search_url, params=params, headers=headers, timeout=15)
            if resp.status_code != 200:
                return ""

            final_url = str(resp.url)
            if "/actresses/" in final_url and final_url != search_url:
                return final_url

            sel = Selector(resp.text)
            href = sel.css('a[href^="/actresses/"]::attr(href)').get()
            if href and href != "/actresses":
                return f"https://www.libredmm.com{href}"
    except Exception:
        pass
    return ""


# ============= TMDB 翻译过滤 =============

_INVALID_TRANSLATIONS = frozenset(
    {
        "Mandarin",
        "Chinese",
        "Japanese",
        "English",
        "Korean",
        "Cantonese",
        "Taiwanese",
        "Min Nan",
        "Hokkien",
    }
)


def _normalize_translation(value: str) -> str:
    normalized = value.strip()
    if normalized in _INVALID_TRANSLATIONS:
        return ""
    return normalized


# ============= Info XML → xlsx 迁移 =============


async def migrate_info_xml_to_xlsx() -> bool:
    """
    将 mapping_info.xml 迁移到 info_database.xlsx。
    迁移成功后返回 True。
    """
    from openpyxl.utils import get_column_letter

    db_path = manager.data_folder / "userdata" / "info_database.xlsx"
    if db_path.exists():
        return True

    xml_path = manager.data_folder / "userdata" / "mapping_info.xml"
    if not xml_path.exists():
        return False

    try:
        parser = None
        try:
            from lxml import etree

            parser = etree.HTMLParser(encoding="utf-8")
        except ImportError:
            pass
        if not parser:
            return False

        content = await _read_text_file(xml_path)
        xml_data = etree.HTML(content.encode("utf-8"), parser)
        info_objects = xml_data.xpath("//a")

        try:
            import openpyxl
        except ImportError:
            LogBuffer.log().write("  ⚠️ [信息映射表] 缺少 openpyxl 库，无法迁移")
            return False

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "信息映射表"
        headers = ["日文名", "中文名", "繁体名", "别名"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = openpyxl.styles.Font(bold=True)
            cell.fill = openpyxl.styles.PatternFill("solid", fgColor="C0C0C0")
            cell.alignment = openpyxl.styles.Alignment(horizontal="center")

        for ob in info_objects:
            jp = ob.get("jp") or ""
            zh_cn = ob.get("zh_cn") or ""
            zh_tw = ob.get("zh_tw") or ""
            keyword = ob.get("keyword") or ""
            ws.append([jp, zh_cn, zh_tw, keyword])

        ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"

        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)) + 2)
                except (TypeError, AttributeError):
                    pass
            adjusted_width = min(max_length, 30)
            ws.column_dimensions[column].width = adjusted_width

        wb.save(db_path)
        wb.close()
        LogBuffer.log().write(f"  ℹ️ [信息映射表] 已从 XML 迁移 {len(info_objects)} 条记录")
        return True
    except Exception as e:
        LogBuffer.log().write(f"  ⚠️ [信息映射表] 迁移失败: {e}")
        return False
