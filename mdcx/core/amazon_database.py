"""
Amazon ASIN 数据库保存功能
用于保存影片番号与 ASIN 对应关系，方便后续统计和复用
"""

import asyncio
from datetime import datetime
from pathlib import Path
from typing import TypedDict


class AsinRecord(TypedDict, total=False):
    """ASIN 记录结构"""

    number: str  # 影片番号
    asin: str  # 亚马逊 ASIN
    product_url: str  # 亚马逊商品详情页链接
    title: str  # 商品标题
    poster_url: str  # 封面图片 URL
    search_keyword: str  # 搜索关键词


def _get_default_excel_path() -> Path:
    """获取默认的 Excel 文件路径，位于 userdata 目录下（与 mapping、watermark 等同目录）"""
    from ..config.manager import manager

    userdata_dir = manager.data_folder / "userdata"
    userdata_dir.mkdir(parents=True, exist_ok=True)
    return userdata_dir / "amazon_asin_database.xlsx"


async def save_asin_to_excel(
    records: list[AsinRecord],
    excel_path: Path | None = None,
    *,
    sheet_name: str = "ASIN 数据库",
    max_rows: int = 100000,
) -> Path:
    """
    保存 ASIN 记录到 Excel 文件

    Args:
        records: ASIN 记录列表
        excel_path: Excel 文件路径，默认保存到运行目录下的 amazon_asin_database.xlsx
        sheet_name: 工作表名称
        max_rows: 单个 sheet 最大行数，超过会自动分 sheet

    Returns:
        Excel 文件路径

    注意：
        需要安装 openpyxl 库：pip install openpyxl
    """
    from ..models.log_buffer import LogBuffer

    try:
        import openpyxl
        from openpyxl.utils import get_column_letter
    except ImportError:
        LogBuffer.log().write("  ⚠️ [ASIN 数据库] 缺少 openpyxl，无法保存 amazon_asin_database.xlsx")
        raise ImportError("请安装 openpyxl 库：pip install openpyxl")

    if excel_path is None:
        excel_path = _get_default_excel_path()
    elif isinstance(excel_path, str):
        excel_path = Path(excel_path)

    excel_path.parent.mkdir(parents=True, exist_ok=True)

    try:
        wb = openpyxl.load_workbook(excel_path)
    except FileNotFoundError:
        wb = openpyxl.Workbook()

    ws = wb.active
    ws.title = sheet_name

    if not ws["A1"].value:
        headers = [
            "影片番号",
            "ASIN 编号",
            "影片链接",
            "商品标题",
            "封面 URL",
            "搜索关键词",
        ]
        # 使用 cell() 直接设置表头，避免 append() 的空行问题
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = openpyxl.styles.Font(bold=True)
            cell.fill = openpyxl.styles.PatternFill("solid", fgColor="C0C0C0")
            cell.alignment = openpyxl.styles.Alignment(horizontal="center")

        ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"

    for record in records:
        row_data = [
            record.get("number", ""),
            record.get("asin", ""),
            record.get("product_url", ""),
            record.get("title", ""),
            record.get("poster_url", ""),
            record.get("search_keyword", ""),
        ]
        ws.append(row_data)

    _format_asin_worksheet(ws)

    try:
        wb.save(excel_path)
        wb.close()
    except PermissionError as e:
        raise PermissionError(f"无法保存 Excel 文件，可能文件正被其他程序打开：{excel_path}") from e

    return excel_path


def _format_asin_worksheet(ws) -> None:
    """格式化 ASIN 数据库工作表：固定表头、自动筛选、列宽、边框、超链接、表头样式。"""
    try:
        import openpyxl
        from openpyxl.utils import get_column_letter

        ws.freeze_panes = "B2"

        last_col = get_column_letter(6)
        ws.auto_filter.ref = f"A1:{last_col}{ws.max_row}"

        header_fill = openpyxl.styles.PatternFill("solid", fgColor="F2F2F2")
        header_font = openpyxl.styles.Font(bold=True, size=11)
        header_align = openpyxl.styles.Alignment(horizontal="center", vertical="center", wrap_text=True)
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_align

        thin = openpyxl.styles.Side(style="thin", color="D0D0D0")
        border = openpyxl.styles.Border(left=thin, right=thin, top=thin, bottom=thin)
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=6):
            for cell in row:
                cell.border = border

        for row in ws.iter_rows(min_row=2, values_only=False):
            for col_idx in [2, 4]:
                cell = row[col_idx]
                val = str(cell.value or "").strip()
                if val and val.startswith("http"):
                    existing_target = cell.hyperlink.target if cell.hyperlink else None
                    if existing_target != val:
                        cell.hyperlink = val
                        cell.style = "Hyperlink"

        caps = {1: 20, 2: 15, 3: 50, 4: 80, 5: 50, 6: 40}
        col_max = [0] * 7
        for row in ws.iter_rows(min_row=2, values_only=True):
            for ci, cell in enumerate(row, 1):
                if cell is None or ci > 6:
                    continue
                s = str(cell)
                width = sum(2 if "\u3040" <= c <= "\u30ff" or "\u4e00" <= c <= "\u9fff" else 1 for c in s)
                col_max[ci] = max(col_max[ci], width)
        for ci in range(1, 7):
            letter = get_column_letter(ci)
            ws.column_dimensions[letter].width = min(col_max[ci] + 2, caps.get(ci, 80))
    except Exception as e:
        from ..models.log_buffer import LogBuffer

        LogBuffer.log().write(f"  ⚠️ [ASIN 数据库] 工作表格式化失败：{e}")


async def save_single_asin_record(
    number: str,
    asin: str,
    title: str = "",
    product_url: str = "",
    poster_url: str = "",
    search_keyword: str = "",
    excel_path: Path | None = None,
) -> bool:
    """
    保存单条 ASIN 记录

    Args:
        number: 影片番号
        asin: ASIN 编号（必须为 10 位字母数字）
        title: 商品标题
        product_url: 亚马逊商品详情页链接
        poster_url: 封面图片 URL
        search_keyword: 搜索关键词
        excel_path: Excel 文件路径

    Returns:
        bool: 保存成功返回 True，失败或跳过返回 False

    示例:
        success = await save_single_asin_record(
            number="ABC-123",
            asin="B0000001",
            title="作品标题",
            product_url="https://www.amazon.co.jp/dp/B0000001",
            poster_url="https://m.media-amazon.com/images/I/xxx.jpg",
        )
        if success:
            print("保存成功")
        else:
            print("保存失败或跳过")
    """
    import re

    if not asin or not asin.strip():
        return False

    asin = asin.strip().upper()

    if not re.match(r"^[A-Z0-9]{10}$", asin):
        return False

    record: AsinRecord = {
        "number": number,
        "asin": asin,
        "product_url": product_url,
        "title": title,
        "poster_url": poster_url,
        "search_keyword": search_keyword,
    }

    try:
        await save_asin_to_excel([record], excel_path)
        return True
    except Exception:
        return False


async def update_asin_record(
    number: str,
    poster_url: str,
    excel_path: Path | None = None,
) -> bool:
    """
    更新已有 ASIN 记录的 poster_url（原地更新，不新增行）

    Args:
        number: 影片番号
        poster_url: 新的封面 URL
        excel_path: Excel 文件路径

    Returns:
        bool: 更新成功返回 True，未找到记录返回 False
    """
    try:
        import openpyxl
    except ImportError:
        return False

    if excel_path is None:
        excel_path = _get_default_excel_path()

    if not excel_path.exists():
        return False

    wb = openpyxl.load_workbook(excel_path)
    ws = wb.active

    updated = False
    for row in ws.iter_rows(min_row=2, values_only=False):
        row_number = str(row[0].value or "").upper()
        if row_number == number.upper():
            row[4].value = poster_url
            updated = True
            break

    if updated:
        wb.save(excel_path)
    wb.close()
    return updated


async def query_asin_database(
    number: str | None = None,
    asin: str | None = None,
    excel_path: Path | None = None,
) -> list[AsinRecord]:
    """
    查询 ASIN 数据库

    Args:
        number: 按番号查询
        asin: 按 ASIN 查询
        excel_path: Excel 文件路径

    Returns:
        匹配的记录列表

    示例:
        results = await query_asin_database(number="ABC-123")
        results = await query_asin_database(asin="B0000001")
    """
    try:
        import openpyxl  # noqa: F401
    except ImportError:
        raise ImportError("请安装 openpyxl 库：pip install openpyxl")

    if excel_path is None:
        excel_path = _get_default_excel_path()

    if not excel_path.exists():
        return []

    try:
        import openpyxl
    except ImportError:
        from ..models.log_buffer import LogBuffer

        LogBuffer.log().write("  ⚠️ [ASIN 数据库] 缺少 openpyxl，无法读取 amazon_asin_database.xlsx")
        return []

    results: list[AsinRecord] = []

    try:
        wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)
        ws = wb.active

        for row_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
            if row_idx == 1:
                continue

            if len(row) < 6:
                continue

            record = AsinRecord(
                number=str(row[0] or ""),
                asin=str(row[1] or ""),
                product_url=str(row[2] or ""),
                title=str(row[3] or ""),
                poster_url=str(row[4] or ""),
                search_keyword=str(row[5] or ""),
            )

            if number and str(record.get("number", "")).upper() == number.upper():
                results.append(record)
            elif asin and str(record.get("asin", "")).upper() == asin.upper():
                results.append(record)

        wb.close()
    except Exception as e:
        from ..models.log_buffer import LogBuffer

        LogBuffer.log().write(f"  ⚠️ [ASIN 数据库] 读取失败：{e}")

    return results


async def export_asin_statistics(
    excel_path: Path | None = None,
    output_path: Path | None = None,
) -> dict:
    """
    导出 ASIN 数据库统计信息

    Returns:
        统计信息字典
    """
    try:
        import openpyxl
    except ImportError:
        raise ImportError("请安装 openpyxl 库：pip install openpyxl")

    if excel_path is None:
        excel_path = _get_default_excel_path()

    if not excel_path.exists():
        return {}

    if output_path is None:
        output_path = excel_path.parent / "amazon_statistics.txt"

    wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)
    ws = wb.active

    total_records = 0

    for row_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if row_idx == 1:
            continue

        if len(row) < 2:
            continue

        total_records += 1

    wb.close()

    stats = {
        "total_records": total_records,
    }

    report = (
        "=" * 60
        + "\n"
        + "Amazon ASIN 数据库统计报告\n"
        + f"生成时间：{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n"
        + "=" * 60
        + "\n\n"
        + f"总记录数：{total_records}\n"
        + "\n"
        + "=" * 60
        + "\n"
    )
    await asyncio.to_thread(output_path.write_text, report, encoding="utf-8")

    return stats
